# -*- coding: utf-8 -*-
"""
Module dedicated to report generation

:Contains:
    :Class:
        - Report

"""

##########################################################################
# Populse_mia - Copyright (C) IRMaGe/CEA, 2018
# Distributed under the terms of the CeCILL license, as published by
# the CEA-CNRS-INRIA. Refer to the LICENSE file or to
# http://www.cecill.info/licences/Licence_CeCILL_V2.1-en.html
# for details.
##########################################################################

# Other import
import json
import os
import platform
import tempfile
import time
from datetime import datetime
from math import floor, log10, modf
from sys import version

# import matplotlib.collections as collections
import matplotlib.pyplot as plt
import nibabel as nib
import numpy as np
import openpyxl
import pandas as pd
import traits.api as traits

# capsul import
from capsul import info as capsul_info
from nilearn.plotting import plot_carpet

# nipype import
from nipype import info as nipype_info

# populse_mia import
from populse_mia import info as mia_info
from populse_mia import sources_images
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch, mm

# reportlab import
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from scipy.interpolate import splev, splrep
from scipy.io import loadmat

# mia_processes import:
from mia_processes import info as mia_processes_info
from mia_processes.utils import (
    PageNumCanvas,
    ReportLine,
    plot_qi2,
    plot_realignment_parameters,
    plot_segmentation,
    plot_slice_planes,
)


class Report:
    """Create pdf report

    IQMs_file --> mriqc individual report (with all IQMs)
    mriqc_group --> mriqc report group
    CVR --> CVR report
    GE2REC --> GE2REC report

    Methods:
      - get_iqms_data
      - co2_inhal_cvr_make_report
      - ge2rec_make_report
      - mriqc_anat_make_report
      - mriqc_func_make_report
      - mriqc_group_make_report

    """

    def __init__(self, report_file, dict4runtime, **kwargs):
        """Create Canvas , create cover and make report"""

        pdfmetrics.registerFont(
            TTFont(
                "DejaVuSans",
                os.path.join(os.path.dirname(__file__), "DejaVuSans.ttf"),
            )
        )
        today_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.dict4runtime = dict4runtime
        ref_exp = "Undefined"

        for key in kwargs:
            setattr(self, key, kwargs[key])

        if "IQMs_file" in kwargs:
            with open(self.IQMs_file) as f:
                self.iqms_data = json.load(f)

            if "anat" in kwargs:
                self.make_report = self.mriqc_anat_make_report
                ref_exp = self.anat
                self.title = (
                    "<font size=18><b>Anatomical Image-Quality "
                    "Metrics summary report</b></font>"
                )

            elif "func" in kwargs:
                self.make_report = self.mriqc_func_make_report
                ref_exp = self.func
                self.title = (
                    "<font size=18><b>Functional Image-Quality "
                    "Metrics summary report</b></font>"
                )

            self.header_title = (
                "<sup rise=20 size=9>$</sup>"
                "<font size=30><b>MRIQ</b></font>"
                "<font size=11>uality</font>"
                "<font size=30><b>C</b></font>"
                "<font size=11>ontrol</font>"
            )
            infos = [
                self.dict4runtime[i]
                for i in (
                    "Site",
                    "Spectro",
                    "StudyName",
                    "AcquisitionDate",
                    "PatientName",
                    "Sex",
                    "Age",
                )
            ]
            infos.insert(4, today_date)
            infos.insert(5, ref_exp)

            headers = [
                "SITE",
                "MRI SCANNER",
                "STUDY NAME",
                "EXAMINATION DATE",
                "MRIQC CALCULATION DATE",
                "NAME OF THE INPUT DATA",
                "PATIENT REFERENCE",
                "PATIENT SEX",
                "PATIENT AGE",
                "SOFTWARES",
            ]

        elif "mriqc_group" in kwargs:
            self.make_report = self.mriqc_group_make_report
            self.title = "<font size=18><b>MRIQc group report</b></font>"

            self.header_title = (
                "<sup rise=20 size=9>$</sup>"
                "<font size=30><b>MRIQ</b></font>"
                "<font size=11>uality</font>"
                "<font size=30><b>C</b></font>"
                "<font size=11>ontrol</font>"
            )
            infos = [today_date]
            headers = [
                "MRIQC GROUPS CALCULATION DATE",
                "SOFTWARES",
            ]

        elif "CVR" in kwargs:
            self.make_report = self.co2_inhal_cvr_make_report
            # ref_exp = {
            #     "norm_anat": self.norm_anat,
            #     "norm_func": self.norm_func,
            # }
            self.title = (
                "<font size=18><b>{0} report: </b>{1}"
                "</font>".format(
                    self.dict4runtime["norm_anat"]["PatientRef"],
                    today_date.split(" ")[0],
                )
            )

            self.header_title = (
                "<font size=30><b>C</b></font>"
                "<font size=11>erebro</font>"
                "<font size=30><b>V</b></font>"
                "<font size=11>ascular</font>"
                "<font size=30><b> R</b></font>"
                "<font size=11>eactivity<br/></font>"
                "<font size=7>assessed using CO<sub>2</sub> inhalation "
                "as a physiological challenge</font>"
            )
            infos = [
                self.dict4runtime["norm_anat"][i]
                for i in (
                    "Site",
                    "Spectro",
                    "StudyName",
                    "AcquisitionDate",
                    "PatientRef",
                    "Sex",
                    "Age",
                    "Pathology",
                )
            ]
            infos.insert(4, today_date)
            # infos.insert(5, ref_exp)
            # Hard-coded, we know the ref data used for IL in the
            # CVR CO2 pipeline:
            infos.insert(10, "CVR_temoins_IL")

            headers = [
                "SITE",
                "MRI SCANNER",
                "STUDY NAME",
                "EXAMINATION DATE",
                "CVR CALCULATION DATE",
                # "NAME OF THE INPUT DATA",
                "PATIENT REFERENCE",
                "PATIENT SEX",
                "PATIENT AGE",
                "PATHOLOGY",
                "REFERENCE GROUP",
                "SOFTWARES",
            ]

        elif "GE2REC" in kwargs:
            self.make_report = self.ge2rec_make_report
            ref_exp = {
                "norm_anat": self.norm_anat,
                "norm_func_gene": self.norm_func_gene,
                "norm_func_reco": self.norm_func_reco,
                "norm_func_recall": self.norm_func_recall,
            }
            self.title = (
                "<font size=18><b>{0} report: </b>{1}"
                "</font>".format(
                    self.dict4runtime["norm_anat"]["PatientName"],
                    today_date.split(" ")[0],
                )
            )

            self.header_title = (
                "<font size=30><b>GE2REC - Langage et " "mémoire</b></font>"
            )
            infos = [
                self.dict4runtime["norm_anat"][i]
                for i in (
                    "Site",
                    "Spectro",
                    "StudyName",
                    "AcquisitionDate",
                    "Sex",
                    "Age",
                    "Pathology",
                    "LateralizationPathology",
                    "DominantHand",
                )
            ]

            headers = [
                "SITE",
                "MRI SCANNER",
                "STUDY NAME",
                "EXAMINATION DATE",
                "PATIENT SEX",
                "PATIENT AGE",
                "PATHOLOGY",
                "LATERALIZATION PATHOLOGY",
                "DOMINANT HAND",
                "SOFTWARES",
            ]

        infos.extend(
            (
                "Python " + version.split()[0],
                "Populse_mia " + mia_info.__version__,
                "Capsul " + capsul_info.__version__,
                "Nipype " + nipype_info.__version__,
                "Mia_processes " + mia_processes_info.__version__,
                "Operating System {0} {1}".format(
                    (
                        "Mac OS X"
                        if platform.system() == "Darwin"
                        else platform.system()
                    ),
                    platform.release(),
                ),
            )
        )

        headers.extend([" "] * 5)
        # Initialises stylesheet with few basic heading and text styles,
        # return a stylesheet object
        self.styles = getSampleStyleSheet()
        # ParagraphStyle gives all the attributes available for formatting
        # paragraphs
        self.styles.add(ParagraphStyle(name="Center", alignment=TA_CENTER))
        self.styles.add(ParagraphStyle(name="Center2", alignment=TA_CENTER))
        # If more than 1 line
        self.styles["Center2"].leading = 24
        self.styles.add(ParagraphStyle(name="Justify", alignment=TA_JUSTIFY))
        self.styles.add(ParagraphStyle(name="Right", alignment=TA_RIGHT))
        self.styles.add(ParagraphStyle(name="Left", alignment=TA_LEFT))
        self.styles.add(ParagraphStyle(name="Left2", alignment=TA_LEFT))
        # For left indent 30
        self.styles["Left2"].leftIndent = 30
        self.styles.add(
            ParagraphStyle(
                name="Bullet1",
                leftIndent=30,
                bulletOffsetY=2,
                bulletIndent=20,
                bulletFontSize=6,
                bulletColor="black",
                bulletText="●",
            )
        )
        self.styles.add(
            ParagraphStyle(
                name="Bullet2",
                leftIndent=60,
                bulletOffsetY=1,
                bulletIndent=50,
                bulletFontSize=6,
                bulletColor="black",
                bulletText="❍",
            )
        )
        # Logo populse
        header_image_1 = os.path.join(
            sources_images.__path__[0], "Logo_populse_square.jpg"
        )
        # 2156px x 2156px
        header_image_1 = Image(header_image_1, 43.0 * mm, 43.0 * mm)
        # Logo Mia
        header_image_2 = os.path.join(
            sources_images.__path__[0], "Logo_populse_mia_HR.jpeg"
        )
        # 2505px x 1843px
        header_image_2 = Image(header_image_2, 54.4 * mm, 40.0 * mm)
        self.image_cov = Table(
            [[header_image_1, header_image_2]], [70 * mm, 70 * mm]
        )
        self.image_cov.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, 0), "LEFT"),
                    ("ALIGN", (-1, 0), (-1, 0), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                ]
            )
        )
        cover_data = [[0, 0]]

        for header, info in zip(headers, infos):
            if isinstance(info, dict):
                info = " AND ".join(str(val) for val in info.values())

            if cover_data == [[0, 0]]:
                if header == " ":
                    cover_data[0][0] = Paragraph(
                        "<para align=right>" + header + "</para>",
                        self.styles["Normal"],
                    )

                else:
                    # Class reportlab.platypus.paragraph.Paragraph
                    # with XML-like markup
                    cover_data[0][0] = Paragraph(
                        "<para align=right><b>" + header + "&nbsp&nbsp&nbsp :"
                        "</b></para>",
                        self.styles["BodyText"],
                    )

                cover_data[0][1] = Paragraph(
                    "<para align=justify>" + info + "</para>",
                    self.styles["Normal"],
                )

            else:
                if header == " ":
                    temp = [
                        Paragraph(
                            "<para align=right>" + header + "</para>",
                            self.styles["Normal"],
                        ),
                        Paragraph(
                            "<para align=justify>" + info + "</para>",
                            self.styles["Normal"],
                        ),
                    ]

                else:
                    temp = [
                        Paragraph(
                            "<para align=right><b>"
                            + header
                            + "&nbsp&nbsp&nbsp :</b></para>",
                            self.styles["BodyText"],
                        ),
                        Paragraph(
                            (
                                "<para align=justify><font size = 5>"
                                + info
                                + "</font></para>"
                                if header == "NAME OF THE INPUT DATA"
                                else "<para align=justify>"
                                + str(info)
                                + "</para>"
                            ),
                            self.styles["Normal"],
                        ),
                    ]

                cover_data.append(temp)

        # colWidths, rowHeights
        self.cover_data = Table(cover_data, [60 * mm, 97 * mm])
        # Careful: Table use (col, raw) and 0-base for top left start as usual
        # OR -1-base for lower right start
        self.cover_data.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                    ("LINEBELOW", (0, -1), (-1, -1), 2, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        self.cover_data.hAlign = "CENTER"

        self.textDisclaimer = (
            "<font size=7 ><i><b>DISCLAIMER</b><br/>Mia "
            "software, from the Populse project, is executed"
            " in a research environment on anonymized data. "
            "The conclusions obtained with Mia software are "
            "an help to diagnose and prognosticate. They do "
            "not substitute themselves to the clinical care "
            "of the physicians and remain under their "
            "responsibilities. Consequently, Populse team is"
            " not responsible for any direct or indirect "
            "damages resulting from the use of data, "
            "informations, or results stemming from the Mia "
            "software. The user recognizes to use these "
            "informations under his sole and exclusive "
            "responsibility.<br/> <br/> <b>DECHARGE DE "
            "RESPONSABILITE</b><br/>Le logiciel Mia, "
            "provenant du projet Populse, est exécuté dans "
            "un environnement de recherche sur des données "
            "anonymisées. Les conclusions obtenues grâce au "
            "logiciel Mia sont une aide au diagnostic et au "
            "pronostic. Elles ne se substituent pas à la "
            "prise en charge médicale des médecins et "
            "demeurent sous leurs responsabilités. Par "
            "conséquent, l'équipe Populse ne peut être tenu "
            "responsable de dommage direct ou indirect "
            "résultant de l'utilisation des données, des "
            "informations ou des résultats issus du logiciel"
            "Mia. L'utilisateur reconnaît utiliser ces "
            "informations sous sa seule et entière "
            "responsabilité.</i> </font>"
        )

        self.neuro_conv = (
            "<font size=9 > <i> 'Neurological' "
            "convention, the left side of the "
            "image corresponds to the left side of "
            "the brain. </i> <br/> </font>"
        )

        # Output document template definition for page; margins and pages size
        self.page = SimpleDocTemplate(
            report_file,
            pagesize=portrait(A4),
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )

        # Canvas creation
        self.report = []

    def get_iqms_data(self, param):
        """Get iqms data"""

        if self.iqms_data.get(param) is None:
            if self.dict4runtime["extra_info"][param][2] is None:
                return Paragraph(
                    "<font size = 11> <b> "
                    + self.dict4runtime["extra_info"][param][0]
                    + " </b> </font>: Not determined",
                    self.styles["Bullet2"],
                )
            else:
                return Paragraph(
                    "<font size = 9><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup></font><font size = 11><b>"
                    + self.dict4runtime["extra_info"][param][0]
                    + "</b></font>: Not determined",
                    self.styles["Bullet2"],
                )

        else:
            if self.dict4runtime["extra_info"][param][2] is None:
                return Paragraph(
                    "<font size = 11> <b>{0}</b> </font>: {1}".format(
                        self.dict4runtime["extra_info"][param][0],
                        (
                            str(
                                round(
                                    self.iqms_data.get(param),
                                    self.dict4runtime["extra_info"][param][1],
                                )
                            )
                            if isinstance(
                                self.dict4runtime["extra_info"][param][1], int
                            )
                            else self.dict4runtime["extra_info"][param][
                                1
                            ].format(self.iqms_data[param])
                        ),
                    ),
                    self.styles["Bullet2"],
                )

            else:
                return Paragraph(
                    "<font size = 9><sup>{0}</sup></font><font size = 11><b>"
                    "{1}</b></font>: {2}".format(
                        self.dict4runtime["extra_info"][param][2],
                        self.dict4runtime["extra_info"][param][0],
                        (
                            str(
                                round(
                                    self.iqms_data.get(param),
                                    self.dict4runtime["extra_info"][param][1],
                                )
                            )
                            if isinstance(
                                self.dict4runtime["extra_info"][param][1], int
                            )
                            else self.dict4runtime["extra_info"][param][
                                1
                            ].format(self.iqms_data[param])
                        ),
                    ),
                    self.styles["Bullet2"],
                )

    def co2_inhal_cvr_make_report(self):
        """Make CVR under CO2 challenge report"""

        # page 1 - cover ##################################################
        ###################################################################

        self.report.append(self.image_cov)
        # width, height
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(Paragraph(self.header_title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 18 * mm))
        self.report.append(Paragraph(self.title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(self.cover_data)
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(self.textDisclaimer, self.styles["Justify"])
        )
        self.report.append(PageBreak())

        # page 2 - Anatomical MRI - Acq & Post-pro parameters############
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b> Anatomical MRI <br/></b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size=15 ><b>Acquisition parameters:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))
        acq_num = self.dict4runtime["norm_anat"]["AcquisitionNumber"]

        if isinstance(acq_num, list):
            acq_num = acq_num[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Protocol name / Acquisition nbr: </b> "
                f"</font> {self.dict4runtime['norm_anat']['ProtocolName']} / "
                f"{acq_num}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Acquisition mode</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Sequence name: </b> "
                f"</font> {self.dict4runtime['norm_anat']['SequenceName']}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Raw geometry parameters</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        d_dim = self.dict4runtime["norm_anat"][
            "Dataset dimensions (Count, X,Y,Z,T...)"
        ]

        if isinstance(d_dim, list):
            d_dim = d_dim[1:]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Dataset dimension:</b> "
                f"</font> {d_dim}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_anat"]["FOV"]

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        if isinstance(fov, list):
            fov = " / ".join(map(str, fov))

        self.report.append(
            Paragraph(
                f"<font size=11><b> FOV (ap / fh / rl) [mm]:</b></font> "
                f"{fov}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        num_sli = self.dict4runtime["norm_anat"]["MaxNumOfSlices"]

        if isinstance(num_sli, list):
            num_sli = num_sli[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Number of slices:</b> "
                f"</font> {num_sli}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_anat"]["SliceThickness"]
        sl_gap = self.dict4runtime["norm_anat"]["SliceGap"]

        if isinstance(sl_thick, list):
            sl_thick = sl_thick[0]

        if isinstance(sl_gap, list):
            sl_gap = sl_gap[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Slice thickness "
                f"/ Slice gap [mm]:</b> "
                f"</font> {sl_thick} / {sl_gap}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        scan_res = self.dict4runtime["norm_anat"]["ScanResolution"]

        if isinstance(scan_res, list):
            scan_res = " / ".join(map(str, scan_res))

        self.report.append(
            Paragraph(
                "<font size=11><b> Scan resolution  (x / y):</b></font> "
                f"{scan_res}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        vox_size = self.dict4runtime["norm_anat"][
            "Grid spacings (X,Y,Z,T,...)"
        ]

        if all(isinstance(elt, (int, float)) for elt in vox_size):
            vox_size = [round(elt, 1) for elt in vox_size]

        if isinstance(vox_size, list):
            vox_size = " / ".join(map(str, vox_size))

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:"
                f"</b> </font> {vox_size}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Other parameters" "</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_anat"]["RepetitionTime"]
        te = self.dict4runtime["norm_anat"]["EchoTime"]
        flipang = self.dict4runtime["norm_anat"]["FlipAngle"]

        if isinstance(flipang, list):
            flipang = round(flipang[0], 1)

        if isinstance(te, list):
            te = round(te[0], 1)

        if isinstance(tr, list):
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        if len(d_dim) == 3:
            dyn_nb = 1

        elif len(d_dim) == 4:
            dyn_nb = d_dim[-1]

        else:
            # TODO: What do we do for 5D. Here if d_dim == "Undefined" or if
            #       d_dim > 4, then dyn_nb = "Undefined"
            dyn_nb = "Undefined"

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Number of dynamics:</b> "
                f"</font> {dyn_nb}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        scan_dur = self.dict4runtime["norm_anat"]["ScanDuration"]

        if isinstance(scan_dur, list):
            scan_dur = scan_dur[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Acquisition duration [s]:</b> "
                f"</font> {scan_dur}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size=15 > <b> Post-processing:<br /> </b> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Spatial processing</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # Hard-coded, we know the state of the normalization in the
        # CVR CO2 pipeline:
        self.report.append(
            Paragraph(
                "<font size=11> <b> Spatial normalization:</b> </font> Y",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        brain_templ = self.dict4runtime["norm_anat"][
            "Affine regularization type"
        ]
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Brain template:"
                f"</b> </font> {brain_templ}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        f_vox_size = self.dict4runtime["norm_anat"]["Voxel sizes"]

        if all(isinstance(elt, (int, float)) for elt in f_vox_size):
            f_vox_size = [round(elt, 1) for elt in f_vox_size]

        if isinstance(f_vox_size, list):
            f_vox_size = " / ".join(map(str, f_vox_size))

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Final voxel size "
                f"(x / y / z) [mm]:</b> "
                f"</font> {f_vox_size}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # Hard-coded, we know the state of the segmentation in the
        # CVR CO2 pipeline:
        self.report.append(
            Paragraph(
                "<font size=11> <b> Segmentation:</b>"
                " </font> Grey matter, white matter, "
                "cerebrospinal fluid, bone, soft tissues",
                self.styles["Bullet2"],
            )
        )
        self.report.append(PageBreak())

        # page 3 - Anatomical MRI - slice planes mosaic display ###########
        ###################################################################
        self.report.append(
            Paragraph(
                "<font size=15 > <b>MNI normalized axial anatomical "
                "images:</b> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 20 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tmpdir = tempfile.TemporaryDirectory()
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1=self.norm_anat_cmap,
            vmin_1=self.norm_anat_vmin,
            vmax_1=self.norm_anat_vmax,
            out_dir=tmpdir.name,
        )
        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 4 - fmri-cvr MRI Acq & Post-pro parameters##############
        ###############################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b> fMRI under a vasoactive stimulus "
                "(hypercapnia)<br/></b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        self.report.append(
            Paragraph(
                "<font size=15 ><b>Acquisition parameters:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))
        acq_num = self.dict4runtime["norm_func"]["AcquisitionNumber"]

        if isinstance(acq_num, list):
            acq_num = acq_num[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Protocol name / Acquisition nbr: "
                f"</b> </font> "
                f"{self.dict4runtime['norm_func']['ProtocolName']} / "
                f"{acq_num}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Acquisition mode</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Sequence name: </b> </font> "
                f"{self.dict4runtime['norm_func']['SequenceName']}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Raw geometry parameters</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        d_dim = self.dict4runtime["norm_func"][
            "Dataset dimensions (Count, X,Y,Z,T...)"
        ]

        if isinstance(d_dim, list):
            d_dim = d_dim[1:]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Dataset dimension:</b> "
                f"</font> {d_dim}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_func"]["FOV"]

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        if isinstance(fov, list):
            fov = " / ".join(map(str, fov))

        self.report.append(
            Paragraph(
                f"<font size=11><b> FOV (ap / fh / rl) [mm]:</b></font> "
                f"{fov}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        num_sli = self.dict4runtime["norm_func"]["MaxNumOfSlices"]

        if isinstance(num_sli, list):
            num_sli = num_sli[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Number of slices:</b> "
                f"</font> {num_sli}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_func"]["SliceThickness"]
        sl_gap = self.dict4runtime["norm_func"]["SliceGap"]

        if isinstance(sl_thick, list):
            sl_thick = sl_thick[0]

        if isinstance(sl_gap, list):
            sl_gap = sl_gap[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Slice thickness "
                f"/ Slice gap [mm]:</b> "
                f"</font> {sl_thick} / {sl_gap}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        scan_res = self.dict4runtime["norm_func"]["ScanResolution"]

        if isinstance(scan_res, list):
            scan_res = " / ".join(map(str, scan_res))

        self.report.append(
            Paragraph(
                "<font size=11><b> Scan resolution  (x / y):</b></font> "
                f"{scan_res}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        raw_func_vox_size_orig = self.dict4runtime["norm_func"][
            "Grid spacings (X,Y,Z,T,...)"
        ]
        if all(
            isinstance(elt, (int, float)) for elt in raw_func_vox_size_orig
        ):
            raw_func_vox_size = [
                round(elt, 1) for elt in raw_func_vox_size_orig
            ]

        if isinstance(raw_func_vox_size, list):
            raw_func_vox_size = " / ".join(map(str, raw_func_vox_size))

        else:
            raw_func_vox_size = "Undefined"

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:</b> "
                f"</font> {raw_func_vox_size} ",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Other parameters" "</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_func"]["RepetitionTime"]
        te = self.dict4runtime["norm_func"]["EchoTime"]
        flipang = self.dict4runtime["norm_func"]["FlipAngle"]

        if isinstance(flipang, list):
            flipang = round(flipang[0], 1)

        if isinstance(te, list):
            te = round(te[0], 1)

        if isinstance(tr, list):
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / Image flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        if len(d_dim) == 3:
            dyn_nb = 1

        elif len(d_dim) == 4:
            dyn_nb = d_dim[-1]

        else:
            # TODO: What do we do for 5D. Here if d_dim == "Undefined" or if
            #       d_dim > 4, then dyn_nb = "Undefined"
            dyn_nb = "Undefined"

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Number of dynamics:</b> "
                f"</font> {dyn_nb}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        scan_dur = self.dict4runtime["norm_func"]["ScanDuration"]

        if isinstance(scan_dur, list):
            scan_dur = scan_dur[0]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Acquisition duration [s]:</b> "
                f"</font> {scan_dur}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Stimulus and performance</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))  # (width, height)
        self.report.append(
            Paragraph(
                "<font size=11> <b>CVR regressor:</b> </font>"
                f"{self.dict4runtime['regressor_physio']['Regressor state']}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b>Stimulation device / Gas name:"
                f"</b> </font> {self.dict4runtime['norm_func']['GasAdmin']} "
                f"/ {self.dict4runtime['norm_func']['Gas']}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # TODO: Stimulus is currently hard-coded at 8%. Should we take the
        #       value from the patient_info dictionary?
        self.report.append(
            Paragraph(
                "<font size=11><b>Stimulus:</b></font> 8% CO<sub>2</sub>",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size=15 > <b> Post-processing:<br /> </b> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Spatial processing</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # Hard-coded, we know the state of the normalization in the
        # CVR CO2 pipeline:
        self.report.append(
            Paragraph(
                "<font size=11> <b> Spatial normalization:</b> </font> Y",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        brain_templ = self.dict4runtime["norm_func"][
            "Affine regularization type"
        ]
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Brain template:"
                f"</b> </font> {brain_templ}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        f_vox_size = self.dict4runtime["norm_func"]["Voxel sizes"]

        if all(isinstance(elt, (int, float)) for elt in f_vox_size):
            f_vox_size = [round(elt, 1) for elt in f_vox_size]

        if isinstance(f_vox_size, list):
            f_vox_size = " / ".join(map(str, f_vox_size))

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size after normalization "
                f"(x / z / y) [mm]:</b> "
                f"</font> {f_vox_size}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fwhm = self.dict4runtime["smooth_norm_func"][
            "FWHM (X, Y, Z) for Smooth"
        ]

        if all(isinstance(elt, (int, float)) for elt in fwhm):
            fwhm = [round(elt, 1) for elt in fwhm]

        if isinstance(fwhm, list):
            fwhm = " / ".join(map(str, fwhm))

        self.report.append(
            Paragraph(
                f"<font size=11> <b>Spatial smoothing "
                f"(fwhm) [mm]:</b> </font> {fwhm}",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                "<font size=11> <b>Head motion as "
                "nuisance regressor in GLM:"
                "</b> </font> Y",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # TODO: The pipeline doesn't currently use ART. Should we implement it?
        self.report.append(
            Paragraph(
                "<font size=11> <b>Artifact detection "
                "tools (ART):</b> </font> N",
                self.styles["Bullet2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # Hard-coded, we know the state of the mask for the func in the
        # CVR CO2 pipeline:
        self.report.append(
            Paragraph(
                "<font size=11> <b> Mask:</b> </font> " "Grey matter",
                self.styles["Bullet2"],
            )
        )
        self.report.append(PageBreak())

        # page 5 - fmri-cvr MRI quality check: movements & EtCO2 regressor ####
        #######################################################################
        sources_images_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "sources_images"
        )
        try:
            data_rp = np.loadtxt(self.realignment_parameters)

        except IOError:
            print(
                "\n==> The "
                + self.realignment_parameters
                + " was not found ! <==\n"
            )
            data_rp = None

        try:
            matDataReg = loadmat(self.regressor_physio)

        except IOError:
            print(
                "\n==> The " + self.regressor_physio + " was not found ! <==\n"
            )
            matDataReg = None

        # TODO: We state that the file to take is "s" + self.norm_func.
        #       1- Since we want the BOLD time course in bold, perhaps
        #          we could take only the self.norm_func?
        #       2- Should we simply add "s" to self.norm_fun or add the
        #          data as an input to the brick?
        folder, file = os.path.split(self.norm_func)
        snorm_func = os.path.join(folder, "s" + file)

        grey_mat_mask = self.mask_003
        try:
            brain_img_snorm_func = nib.load(snorm_func)
            snorm_func_data = brain_img_snorm_func.get_fdata()
            brain_img_grey_mat_mask = nib.load(grey_mat_mask)
            mask_data = brain_img_grey_mat_mask.get_fdata()
            binary_mask = np.where(mask_data > 0.1, 1, 0)
            masked_data = snorm_func_data * binary_mask[..., np.newaxis]
            bold_signal_tc = np.mean(masked_data, axis=(0, 1, 2))

        except Exception:
            print("\n==> The " + grey_mat_mask + " was not found ! <==\n")
            bold_signal_tc = None

        im_qualCheck = None
        im_qualCheckTra = None
        im_qualCheckRot = None
        im_qualCheckReg = None
        im_qualCheckBoldTC = None

        if data_rp is not None:
            out_file_tra = os.path.join(
                tmpdir.name,
                self.dict4runtime["norm_anat"]["PatientRef"]
                + "_CVR_QualityControlMeasure_translation.png",
            )
            out_file_rot = os.path.join(
                tmpdir.name,
                self.dict4runtime["norm_anat"]["PatientRef"]
                + "_CVR_QualityControlMeasure_Rotation.png",
            )
            qc = plot_realignment_parameters(
                self.realignment_parameters,
                raw_func_vox_size_orig,
                out_file_tra,
                out_file_rot,
            )
            if qc == 0:
                # 912px × 892px
                im_qualCheck = Image(
                    os.path.join(sources_images_dir, "No-check-mark.png"),
                    13.0 * mm,
                    12.7 * mm,
                )
            elif qc == 1:
                # 940px × 893px
                im_qualCheck = Image(
                    os.path.join(sources_images_dir, "OK-check-mark.png"),
                    13.0 * mm,
                    12.4 * mm,
                )
            im_qualCheckTra = Image(out_file_tra, 6.468 * inch, 3.018 * inch)
            im_qualCheckRot = Image(out_file_rot, 6.468 * inch, 3.018 * inch)

        # figsize in inches
        fig = plt.figure(figsize=(12, 5.6), facecolor="white")
        # fig.set_size_inches(254 / 25.4, 142 / 25.4)
        ax = fig.add_subplot(111)
        fig.subplots_adjust(
            left=None,
            bottom=None,
            right=None,
            top=None,
            wspace=None,
            hspace=None,
        )
        xLim = None
        if matDataReg is not None:

            if xLim is None:
                xLim = len(matDataReg["R"])

            ax.clear()
            ax.set_title(
                r"$\mathsf{EtCO_2}$ variation regressor "
                f"("
                f"{self.dict4runtime['regressor_physio']['Regressor state']}"
                f")",
                fontsize=20,
                y=1.03,
            )
            ax.set_xlabel("Dynamic scans", fontsize=14)
            ax.set_ylabel(r"$\mathsf{\Delta EtCO_2}$ ( mmHg)", fontsize=14)
            ax.plot(matDataReg["R"], label=r"$\mathsf{EtCO_2}$")
            # ax.legend(loc='best')
            ax.set_xlim(0, xLim)
            ax.set_ylim(
                matDataReg["R"].min() * 1.1, matDataReg["R"].max() * +1.1
            )
            ax.yaxis.grid(
                True, linestyle="-", which="major", color="grey", alpha=0.5
            )
            ax.xaxis.grid(
                True, linestyle="-", which="major", color="grey", alpha=0.5
            )
            out_file_reg = os.path.join(
                tmpdir.name,
                self.dict4runtime["norm_anat"]["PatientRef"]
                + "_CVR_QualityControlMeasure_EtCO2.png",
            )
            # High resolution: 2000px × 1118px.
            fig.savefig(out_file_reg, format="png", dpi=200)
            im_qualCheckReg = Image(out_file_reg, 7.187 * inch, 3.354 * inch)

        if bold_signal_tc is not None:

            if xLim is None:
                xLim = brain_img_snorm_func.shape[3]

            ax.clear()
            ax.set_title(
                "BOLD signal timecourse (smoothed)", fontsize=20, y=1.03
            )
            ax.set_xlabel("Dynamic scans", fontsize=14)
            ax.set_ylabel("BOLD response (A. U.)", fontsize=14)
            # With smoothing
            x_sparse = np.linspace(
                0, len(bold_signal_tc) - 1, len(bold_signal_tc)
            )
            x_dense = np.linspace(
                0, len(bold_signal_tc) - 1, len(bold_signal_tc) * 10
            )

            ix = np.mean(bold_signal_tc)
            #  Lagrange polynomial interpolation coefficient determined using
            #  the experimental average bold_signal_tc (ix) and the best s (y)
            #  giving a good visual result when smoothing the bold_signal_tc.
            s = (
                (1.64e-24) * ix**5
                - (6.729e-17) * ix**4
                + (2.49e-10) * ix**3
                + (6.301e-05) * ix**2
                + (0.03408) * ix
                - 1.249
            )
            tck = splrep(x_sparse, bold_signal_tc, s=s)
            ax.plot(x_dense, splev(x_dense, tck, der=0), label="BOLD response")
            # ax.legend(loc="best")
            ax.set_xlim(0, xLim)
            ax.set_ylim(
                bold_signal_tc.min() * 0.997, bold_signal_tc.max() * +1.003
            )
            ax.yaxis.grid(
                True, linestyle="-", which="major", color="grey", alpha=0.5
            )
            ax.xaxis.grid(
                True, linestyle="-", which="major", color="grey", alpha=0.5
            )
            out_file_rot = os.path.join(
                tmpdir.name,
                self.dict4runtime["norm_anat"]["PatientRef"]
                + "_CVR_QualityControlMeasure_BOLDtimeCourse.png",
            )
            # High resolution: 2000px × 1118px.
            fig.savefig(out_file_rot, format="png", dpi=200)
            im_qualCheckBoldTC = Image(
                out_file_rot, 7.187 * inch, 3.354 * inch
            )

        qualCheckMess = Paragraph(
            "<font size=14 > <b> fMRI quality "
            "check: movements </b> </font>",
            self.styles["Left"],
        )

        if im_qualCheck is None:
            im_qualCheck = Paragraph(
                "<font size=8 > Automatic evaluation not available </font>",
                self.styles["Center"],
            )

        im_qualCheck.hAlign = "CENTER"
        title = [[qualCheckMess, im_qualCheck]]
        t = Table(title, [110 * mm, 30 * mm])  # colWidths, rowHeight
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        t.hAlign = "LEFT"
        self.report.append(t)

        if im_qualCheckTra is None:
            im_qualCheckTra = Paragraph(
                "<font size=14 > Linear head motion parameters not "
                "available </font>",
                self.styles["Center"],
            )
            self.report.append(Spacer(0 * mm, 45 * mm))

        im_qualCheckTra.hAlign = "CENTER"
        self.report.append(im_qualCheckTra)

        if im_qualCheckRot is None:
            im_qualCheckRot = Paragraph(
                "<font size=14 > Rotational head motion parameters "
                "not available </font>",
                self.styles["Center"],
            )
            self.report.append(Spacer(0 * mm, 45 * mm))

        im_qualCheckRot.hAlign = "CENTER"
        self.report.append(im_qualCheckRot)

        if data_rp is None:
            self.report.append(Spacer(0 * mm, 45 * mm))

        if im_qualCheckReg is None:
            im_qualCheckReg = Paragraph(
                "<font size=14 > EtCO<sub>2</sub> variation regressor "
                "parameters not available </font>",
                self.styles["Center"],
            )

            if data_rp is not None:
                self.report.append(Spacer(0 * mm, 35 * mm))

        im_qualCheckReg.hAlign = "CENTER"
        self.report.append(im_qualCheckReg)
        self.report.append(PageBreak())

        # page 6 - fmri-cvr MRI quality check: BOLD signal timecourse Vs EtCO2
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=14 > <b> fMRI quality "
                "check: BOLD signal timecourse Vs "
                "EtCO<sub>2</sub> model </b> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 30 * mm))

        if im_qualCheckBoldTC is None:
            im_qualCheckBoldTC = Paragraph(
                "<font size=14 > BOLD signal time course not "
                "available </font>",
                self.styles["Center"],
            )
            self.report.append(Spacer(0 * mm, 20 * mm))

        im_qualCheckBoldTC.hAlign = "CENTER"
        self.report.append(im_qualCheckBoldTC)
        self.report.append(Spacer(0 * mm, 30 * mm))

        if matDataReg is None:
            self.report.append(Spacer(0 * mm, 20 * mm))

        self.report.append(im_qualCheckReg)
        self.report.append(PageBreak())

        # page 7 - BOLD: MNI normalized axial images, 1st dyn ################
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=14 > <b> BOLD: MNI "
                "normalized axial images (1<sup>st</sup> "
                "dynamic)</b> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 20 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_func,
            fig_rows=self.norm_func_fig_rows,
            fig_cols=self.norm_func_fig_cols,
            slice_start=self.norm_func_inf_slice_start,
            slice_step=self.norm_func_slices_gap,
            cmap_1=self.norm_func_cmap,
            vmin_1=self.norm_func_vmin,
            vmax_1=self.norm_func_vmax,
            out_dir=tmpdir.name,
        )
        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # page 8 - parametric maps: beta / EtCO2 #############################
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=14><b> Parametric maps: </b></font>"
                "<font name='DejaVuSans' size=14><b>β</b></font>"
                "<font size=14><b> weight values in </b></font>"
                "<font name='DejaVuSans' size=14><b>Δ</b></font>"
                "<font size=14><b>(%BOLD) / EtCO<sub>2</sub> "
                "(mmHg)</b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 20 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.beta_image,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1=self.norm_anat_cmap,
            vmin_1=self.norm_anat_vmin,
            vmax_1=self.norm_anat_vmax,
            cmap_2=self.beta_cmap,
            vmin_2=self.beta_vmin,
            vmax_2=self.beta_vmax,
            out_dir=tmpdir.name,
        )

        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # page 9 - parametric maps: SPMt #####################################
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=14><b> Parametric maps: statistic parametric "
                "t-Map (SPMt) </b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 20 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_image,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1=self.norm_anat_cmap,
            vmin_1=self.norm_anat_vmin,
            vmax_1=self.norm_anat_vmax,
            cmap_2=self.spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )

        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 10 - Laterality indexes #######################################
        ######################################################################
        self.report.append(
            Paragraph(
                f"<font size=16> <b> Laterality index values for "
                f"{self.dict4runtime['norm_anat']['PatientRef']} "
                f"against the box and whisker plot for a reference "
                f"population <br/> </b> </font>",
                self.styles["Center2"],
            )
        )
        self.report.append(Spacer(0 * mm, 0 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        CVR_ref_data = self.dict4runtime["CVR_ref_data"]
        # Sheets used in the CVR_ref_data file for boxplot
        sheetsRef = ["CVR_temoins_IL"]
        bookRef = openpyxl.load_workbook(CVR_ref_data)
        sheet = bookRef[sheetsRef[0]]
        # Number of the first column used for ROI data (in sheet)
        colStart = 8
        # List of ROIs in the CVR_temoins_IL sheet order, of the
        # CVR_ref_pop_SEMVIE-16.xlsx file
        xticklab = []

        for col_index in range(colStart, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col_index).value
            xticklab.append(value)

        # Reference dictionary of laterality indexes for each ROI
        # IL_Ref = {'ROI_STR': [-0.0079, -0.01, ...], 'ACA': ...}
        IL_Ref = {}
        col_index = colStart - 1

        for roi in xticklab:
            col_index += 1
            val_list = []

            for row_index in range(1, sheet.max_row + 1):

                if row_index == 1:
                    continue

                value = sheet.cell(row=row_index, column=col_index).value

                if not isinstance(value, (int, float)):
                    continue

                else:
                    val_list.append(value)

            IL_Ref[roi] = val_list

        # Laterality index for each ROI of the patient
        # Patient dictionary of laterality index for each ROI
        # IL_Pat = {'ACA': -0.0041, 'ACM': ...}
        res_anal_data = os.path.join(
            self.output_directory,
            self.dict4runtime["norm_anat"]["PatientName"] + "_data",
            "results_aggregation",
            "BOLD_IL_mean_beta.xls",
        )

        max_timeout = 250  # Max timeout in seconds
        start_time = time.time()

        # TODO: This is a 2-cts hack in case res_anal_data doesn't already
        #       exist. We can also add this data as input. In this case,
        #       the brick will wait until the files exist.
        #       This would be cleaner.
        i = 0
        while not os.path.exists(res_anal_data):

            if time.time() - start_time > max_timeout:
                print(
                    f"CVR make report: Max timeout "
                    f"reached ({max_timeout}s). The "
                    f"file does not exist."
                )
                break

            print(
                f"{res_anal_data} file does not exist yet. "
                f"Waiting...{i} {time.time() - start_time}s"
            )
            time.sleep(1)
            i += 1

        with open(res_anal_data, "r") as data:
            r_data = [elt.replace("beta_", "") for elt in data.readlines()]

        # IL in ROIs, uses \t for split() to deal with space
        # fmt: off
        IL_Pat = {
            il: float(val)
            for il, val in zip(
                r_data[0].split("\t")[colStart - 1:-1],
                r_data[1].split("\t")[colStart - 1:-1],
            )
        }
        # Data on patient and experiment
        # dat_Pat = {'subjects': 'alej170316', 'patho': ...}
        dat_Pat = {
            dat: val
            for dat, val in zip(
                r_data[0].split("\t")[0:colStart - 1],
                r_data[1].split("\t")[0:colStart - 1],
            )
        }
        # fmt: on
        fig.set_size_inches(15, 15)
        ax.clear()
        ax.yaxis.grid(
            True, linestyle="-", which="major", color="grey", alpha=0.5
        )
        ax.xaxis.grid(
            True, linestyle="-", which="major", color="grey", alpha=0.5
        )
        ax.set_axisbelow(True)
        ax.set_title(
            "Laterality Index from BOLD functional MRI study of the "
            "\n cerebral vasoreactivity to hypercapnia",
            fontsize=23,
            y=1.01,
        )
        ax.set_xlabel("Regions Of Interest", fontsize=18)
        ax.set_ylabel(
            r"$ \mathsf{Laterality \hspace{0.5} Index \hspace{0.5} "
            r"= \hspace{0.5} ( \beta_{left} \hspace{0.5} - "
            r"\hspace{0.5} \beta_{right} ) \hspace{0.5} / "
            r"\hspace{0.5} ( \beta_{left} \hspace{0.5} + "
            r"\hspace{0.5} \beta_{right})} $",
            fontsize=20,
        )
        roi_index = 0

        for roi in xticklab:
            roi_index += 1
            ax.boxplot(
                IL_Ref[roi],
                positions=[roi_index] * len(sheetsRef),
                widths=0.6,
                showfliers=False,
            )

        # TODO: this is a quick patch to correct the nomenclature difference
        #       between ref and pat.
        #       The change must be made in the brick producing the results:
        #       Ex. Currently beta_ROI-FRON, whereas beta_ROI_FRON is
        #       necessary to synchronize with the reference nomenclature.
        IL_Pat_list = {k.replace("-", "_"): v for k, v in IL_Pat.items()}
        IL_Pat_list = [IL_Pat_list[c] for c in xticklab if c in IL_Pat_list]

        if len(IL_Pat_list) != len(xticklab):
            print(
                "\nReportCO2inhalCvr brick warning: the number of values is "
                "not the same for patient and reference when automatically "
                "generating the graph comparing patient and reference...\n"
            )

        ax.plot(
            np.linspace(1, len(IL_Pat_list), len(IL_Pat_list), endpoint=True),
            IL_Pat_list,
            "bo-",
        )

        # Adjust tick labels
        for label in ax.yaxis.get_majorticklabels():
            label.set_fontsize(14)

        # Set x-axis properties
        ax.set_xticklabels(xticklab, rotation=45, fontsize=14)
        ax.set_xticks(range(1, len(xticklab) + 1))
        ax.set_xlim(0, len(xticklab) + 1)
        ax.get_xaxis().set_tick_params(direction="out")
        # Set y-axis properties
        ax.set_ylim(auto=True)
        ax.get_yaxis().set_tick_params(direction="out")
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        # RGB color chartreuse
        # BrokenBarHCollection is deprecated since matplotlib 3.7
        # c1 = collections.BrokenBarHCollection(
        #     [xlim],
        #     (0, ylim[0]),
        #     facecolor=(127.0 / 255.0, 255.0 / 255.0, 0),
        #     alpha=0.2,
        # )
        ax.fill_between(
            np.linspace(xlim[0], xlim[1], 3),
            0,
            ylim[0],
            facecolor=(127.0 / 255.0, 255.0 / 255.0, 0),
            alpha=0.2,
        )
        # RGB color lavender
        # c2 = collections.BrokenBarHCollection(
        #     [xlim],
        #     (0, ylim[1]),
        #     facecolor=(230.0 / 255.0, 230.0 / 255.0, 250.0 / 255.0),
        #     alpha=0.5,
        # )
        ax.fill_between(
            np.linspace(xlim[0], xlim[1], 3),
            0,
            ylim[1],
            facecolor=(230.0 / 255.0, 230.0 / 255.0, 250.0 / 255.0),
            alpha=0.5,
        )
        # ax.add_collection(c1)
        # ax.add_collection(c2)
        ax.set_ylim(ylim)

        for i, j in enumerate(IL_Pat_list, start=1):

            try:
                j = round(j, -int(floor(log10(abs(modf(j)[0])))) + 1)

            except ValueError:
                pass

            if (j - (0.02 * abs(ylim[0] - ylim[1])) > ylim[0]) and (
                j - (0.02 * abs(ylim[0] - ylim[1])) < ylim[1]
            ):
                ax.annotate(
                    str(j),
                    xy=(
                        i + (0.02 * (xlim[1] - xlim[0])),
                        j - (0.02 * abs(ylim[0] - ylim[1])),
                    ),
                )
            else:
                ax.annotate(str(j), xy=(i + 0.1, j))

        ax.text(
            0.01,
            0.99,
            "Right impairment",
            fontsize=26,
            color="red",
            backgroundcolor="none",
            horizontalalignment="left",
            verticalalignment="top",
            transform=ax.transAxes,
        )

        ax.text(
            0.01,
            0.01,
            "Left impairment",
            fontsize=26,
            color="red",
            backgroundcolor="none",
            horizontalalignment="left",
            verticalalignment="bottom",
            transform=ax.transAxes,
        )

        (leg_bp,) = ax.plot([0, 0], "k-")
        (leg_curv,) = ax.plot([0, 0], "bo-")
        ax.legend(
            (leg_bp, leg_curv),
            (
                sheetsRef[0],
                dat_Pat["subjects"] + " (" + dat_Pat["patho"] + ")",
            ),
        )
        leg_bp.set_visible(False)
        leg_curv.set_visible(False)
        ax.plot([0, len(xticklab) + 1], [0, 0], "k-")
        # width, height. in inches
        fig.set_size_inches(400 / 25.4, 260 / 25.4)  # width, height. in inches
        out_file_reg = os.path.join(
            tmpdir.name,
            dat_Pat["subjects"] + "_ILvaso-Patient.png",
        )
        fig.savefig(out_file_reg, dpi=200, format="png", bbox_inches="tight")
        im_PatVsBoxPlot = Image(out_file_reg, 175.0 * mm, 138.5 * mm)
        im_PatVsBoxPlot.hAlign = "CENTER"
        self.report.append(Spacer(0 * mm, 40 * mm))
        self.report.append(im_PatVsBoxPlot)
        self.report.append(PageBreak())

        # page 11 - Vascular territories used for IL #########################
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=14><b> Vascular territories of the cerebral "
                "arteries used for IL determination </b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 15 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        conv_roi_dir = os.path.join(
            self.output_directory,
            self.dict4runtime["norm_anat"]["PatientName"] + "_data",
            "ROI_data",
            "convROI_BOLD",
        )
        arter_terr = ["convACA", "convACM", "convACP", "convPICA", "convSCA"]
        arter_terr_files = [
            os.path.join(conv_roi_dir, f"{a_t}_{l_r}.nii")
            for a_t in arter_terr
            for l_r in ["L", "R"]
        ]

        max_timeout = 60  # Max timeout in seconds
        start_time = time.time()

        # TODO: This is a 2-cts hack in case arter_terr_files doesn't already
        #       exist. We can also add this data as input. In this case,
        #       the brick will wait until the files exist.
        #       This would be cleaner
        for elmt in arter_terr_files:
            i = 0

            while not os.path.exists(elmt):

                if time.time() - start_time > max_timeout:
                    print(
                        f"CVR make report: Max timeout "
                        f"reached ({max_timeout}s). The "
                        f"file does not exist."
                    )
                    break

                print(
                    f"{elmt} file does not exist yet. "
                    f"Waiting...{i} {time.time() - start_time}s"
                )
                time.sleep(1)
                i += 1

        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=arter_terr_files,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1=self.norm_anat_cmap,
            vmin_1=self.norm_anat_vmin,
            vmax_1=self.norm_anat_vmax,
            cmap_2=[
                (255, 0, 0),  # red
                (0, 255, 0),  # lime (green)
                (0, 0, 255),  # blue
                (210, 105, 30),  # chocolate
                (255, 255, 0),  # yellow
            ],
            vmin_2=None,
            vmax_2=None,
            out_dir=tmpdir.name,
            out_name="arterialTerritories_axial",
        )

        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        footnote_artery = [
            [
                Paragraph(
                    "<b> "
                    '<font size = 12 color = "rgb(0, 64, 255)"> '
                    "Posterior cerebral artery </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(255, 64, 64)"> '
                    "Anterior cerebral artery </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(0, 255, 64)"> '
                    "Middle cerebral artery </font> <br/>"
                    '<font size = 12 color = "rgb(210, 105, 30)"> '
                    "Posterior inferior cerebellar artery </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(228, 228, 28)"> '
                    "Superior cerebellar artery </font> "
                    "</b>",
                    self.styles["Center"],
                )
            ]
        ]
        data = Table(footnote_artery, [7.4803 * inch])
        data.setStyle(
            TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.black)])
        )
        data.hAlign = "CENTER"
        self.report.append(data)

        self.report.append(PageBreak())

        # page 12 -  Lobes of the brain used for IL ##########################
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=16><b> Lobes of the brain used for IL "
                "determination </b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 15 * mm))
        self.report.append(
            Paragraph(
                '<font size=9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        brain_lobes = [
            "convROI-CING",
            "convROI-FRON",
            "convROI-INSULA",
            "convROI-OCC",
            "convROI-PAR",
            "convROI-STR",
            "convROI-TEMP",
            "convROI-THA",
        ]
        brain_lobes_files = [
            os.path.join(conv_roi_dir, f"{b_l}_{l_r}.nii")
            for b_l in brain_lobes
            for l_r in ["L", "R"]
        ]
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=brain_lobes_files,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1=self.norm_anat_cmap,
            vmin_1=self.norm_anat_vmin,
            vmax_1=self.norm_anat_vmax,
            cmap_2=[
                (0, 255, 255),  # cyan (aqua)
                (255, 255, 0),  # yellow
                (210, 105, 30),  # chocolate
                (0, 0, 255),  # blue
                (255, 0, 0),  # red
                (255, 165, 0),  # moccasin
                (0, 255, 0),  # lime (green)
                (255, 0, 255),  # magenta (fuschia)
            ],
            # cmap_2=[
            #     (0, 255, 255),  # cyan (aqua)
            #     (255, 255, 0),  # yellow
            #     # (210, 105, 30),   # chocolate
            #     (128, 0, 128),  # purple
            #     (0, 0, 255),  # blue
            #     (255, 0, 0),  # red
            #     # (255, 228, 181),  # moccasin
            #     (255, 140, 0),  # orange
            #     (0, 255, 0),  # lime (green)
            #     (255, 0, 255),  # magenta (fuschia)
            # ],
            vmin_2=None,
            vmax_2=None,
            out_dir=tmpdir.name,
            out_name="roisTerritories_axial",
        )
        # remainder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        footnote_artery = [
            [
                Paragraph(
                    "<b> "
                    '<font size = 12 color = "rgb(0, 64, 255)"> '
                    "Occipital lobe </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(255, 64, 64)"> '
                    "Parietal lobe </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(0, 255, 64)"> '
                    "Temporal lobe </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(210, 105, 30)"> '
                    "Insular lobe </font> <br/>"
                    '<font size = 12 color = "rgb(228, 228, 28)"> '
                    "Frontal lobe </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(0, 255, 255)"> '
                    "Cingulate cortex </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(255, 64, 255)"> '
                    "Thalamus </font> "
                    '<font size = 12 color = "grey"> - </font>'
                    '<font size = 12 color = "rgb(255, 228, 181)"> '
                    "Striatum </font> "
                    "</b>",
                    self.styles["Center"],
                )
            ]
        ]
        data = Table(footnote_artery, [7.4803 * inch])
        data.setStyle(
            TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.black)])
        )
        data.hAlign = "CENTER"
        self.report.append(data)

        self.page.build(self.report, canvasmaker=PageNumCanvas)
        tmpdir.cleanup()

    def ge2rec_make_report(self):
        """Make GE2REC report"""
        norm_func_cmap = "Greys_r"
        norm_anat_cmap = "Greys_r"
        spmT_cmap = "rainbow"
        # axial fig
        norm_anat_fig_rows = 3
        norm_anat_fig_cols = 5
        norm_anat_inf_slice_start = 55
        norm_anat_slices_gap = 5
        norm_func_fig_rows = 4
        norm_func_fig_cols = 5
        norm_func_inf_slice_start = 10
        norm_func_slices_gap = 2
        # sag fig
        norm_anat_fig_rows_sag = 2
        norm_anat_fig_cols_sag = 5
        norm_anat_inf_slice_start_sag = 35
        norm_anat_slices_gap_sag = 10
        # c0r fig
        norm_anat_fig_rows_cor = 1
        norm_anat_fig_cols_cor = 5
        norm_anat_inf_slice_start_cor = 60
        norm_anat_slices_gap_cor = 15

        # page 1 - cover ##################################################
        ###################################################################

        self.report.append(self.image_cov)
        # width, height
        self.report.append(Spacer(0 * mm, 8 * mm))
        self.report.append(Paragraph(self.header_title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(Paragraph(self.title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(self.cover_data)
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(self.textDisclaimer, self.styles["Justify"])
        )
        self.report.append(PageBreak())

        # page 2 - Generation task############
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>Production langage</b> - "
                "Carte Statistique<br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=10 >"
                "<b>Activation cérébrales pour la tâche de génération de "
                "phrases à partir de mots entendus</b>:"
                "</font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 3 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tmpdir = tempfile.TemporaryDirectory()
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_gene,
            fig_rows=norm_anat_fig_rows,
            fig_cols=norm_anat_fig_cols,
            slice_start=norm_anat_inf_slice_start,
            slice_step=norm_anat_slices_gap,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=127 * mm)
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(Spacer(0 * mm, 4 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_gene,
            plane="sag",
            fig_rows=norm_anat_fig_rows_sag,
            fig_cols=norm_anat_fig_cols_sag,
            slice_start=norm_anat_inf_slice_start_sag,
            slice_step=norm_anat_slices_gap_sag,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=95 * mm)
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 3 - Generation task############
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>Production langage</b> - "
                "Evaluation quantitative<br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=10 > Représentation de la latéralisation de la "
                "production du langage pour la région frontale et la région "
                "temporal. <br/> Les courbes noires montrent l'évolution de "
                "l'index de latéralité en fonction du seuil statistique. "
                "<br/> <b>N.B: </b> <i>Plus le seuil est élevé moins il "
                "existe de faux positifs. Il est fortement conseillé de "
                "prendre en compte un seuil supérieur à 3. </i> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        for li_curve in self.li_curves:
            im = Image(li_curve, 6.468 * inch, 3.018 * inch)
            im.hAlign = "CENTER"
            self.report.append(im)
            self.report.append(Spacer(0 * mm, 5 * mm))

        self.report.append(Spacer(0 * mm, 15 * mm))
        self.report.append(
            Paragraph(
                "<font size=10 > <i> <b>Réference</b>: LI toolbox - "
                "Wilke, Marko, and Karen Lidzba. "
                "“LI-Tool: A New Toolbox to Assess Lateralization in "
                "Functional MR-Data.” Journal of Neuroscience Methods "
                "163, no. 1 (June 15, 2007): 128–36. "
                "https://doi.org/10.1016/j.jneumeth.2007.01.026.<br/>"
                "Veuillez noter que ce logiciel n'a pas de marquage "
                "CE ou FDA. </i> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))

        self.report.append(PageBreak())

        # page 4 - Generation task with memory############
        #################################################################
        if self.spmT_gene_enco:
            if self.spmT_gene_enco not in ["<undefined>", traits.Undefined]:
                self.report.append(
                    Paragraph(
                        "<font size=18 ><b>Mémoire encodage implicite</b> - "
                        "Carte Statistique<br/></font>",
                        self.styles["Left"],
                    )
                )
                self.report.append(Spacer(0 * mm, 4 * mm))
                self.report.append(
                    Paragraph(
                        "<font size=10 >"
                        "<b>Activations cérébrales pour l'encodage implicite "
                        "</b> (obtenues à partir de la tâche de génération "
                        "en prenant en compte seulement les mots dont le "
                        "sujet se souvent lors de la tâche de </font>",
                        self.styles["Left"],
                    )
                )
                self.report.append(Spacer(0 * mm, 4 * mm))
                slices_image = plot_slice_planes(
                    data_1=self.norm_anat,
                    data_2=self.spmT_gene_enco,
                    fig_rows=norm_anat_fig_rows,
                    fig_cols=norm_anat_fig_cols,
                    slice_start=norm_anat_inf_slice_start,
                    slice_step=norm_anat_slices_gap,
                    cmap_1=norm_anat_cmap,
                    cmap_2=spmT_cmap,
                    vmin_2=self.spmT_vmin,
                    vmax_2=self.spmT_vmax,
                    out_dir=tmpdir.name,
                )
                slices_image = Image(
                    slices_image, width=177 * mm, height=127 * mm
                )
                slices_image.hAlign = "CENTER"
                self.report.append(slices_image)
                self.report.append(Spacer(0 * mm, 4 * mm))
                slices_image = plot_slice_planes(
                    data_1=self.norm_anat,
                    data_2=self.spmT_gene_enco,
                    plane="cor",
                    fig_rows=norm_anat_fig_rows_cor,
                    fig_cols=norm_anat_fig_cols_cor,
                    slice_start=norm_anat_inf_slice_start_cor,
                    slice_step=norm_anat_slices_gap_cor,
                    cmap_1=norm_anat_cmap,
                    cmap_2=spmT_cmap,
                    vmin_2=self.spmT_vmin,
                    vmax_2=self.spmT_vmax,
                    out_dir=tmpdir.name,
                )
                slices_image = Image(
                    slices_image, width=177 * mm, height=50 * mm
                )
                slices_image.hAlign = "CENTER"
                self.report.append(slices_image)

                self.report.append(Spacer(0 * mm, 4 * mm))

                self.report.append(PageBreak())

        # page 5 - Recognition task############
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>Mémoire reconnaissance</b> "
                "- Carte Statistique <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        if self.correct_response:
            if self.correct_response not in ["<undefined>", traits.Undefined]:
                m1 = Paragraph(
                    "<font size=10 >"
                    "<b>Performances comportementales </b> "
                    "(mesurées pendant la tâche de reconnaisance "
                    "des mots entendus à la tâche précédente à partir d'image)"
                    "</font>",
                    self.styles["Left"],
                )

                m2 = Paragraph("<font size=10 ><b>% CR</b></font>")
                m3 = Paragraph("<font size=10 ><b>% erreur</b></font>")
                m4 = Paragraph("<font size=10 ><b>anciens mots</b></font>")
                m5 = Paragraph("<font size=10 ><b>nouveaux mots</b></font>")
                df = pd.read_csv(self.correct_response)

                behavioural_data = [
                    [m1, "", m2, m3],
                    ["", m4, df["correct_old"][0], df["error_old"][0]],
                    ["", m5, df["correct_new"][0], df["error_new"][0]],
                ]
                t = Table(behavioural_data)
                t.setStyle(
                    TableStyle(
                        [
                            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                            ("SPAN", (0, 0), (0, -1)),
                        ]
                    )
                )
                t.hAlign = "LEFT"
                self.report.append(t)
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=10 >"
                " <b>Activation cérébrales pour la tâche de mémoire</b>"
                " (reconnaisance des mots entendus à la tâche précédente "
                "à partir d'image):</font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_reco,
            fig_rows=norm_anat_fig_rows,
            fig_cols=norm_anat_fig_cols,
            slice_start=norm_anat_inf_slice_start,
            slice_step=norm_anat_slices_gap,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=127 * mm)
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(Spacer(0 * mm, 4 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_reco,
            plane="cor",
            fig_rows=norm_anat_fig_rows_cor,
            fig_cols=norm_anat_fig_cols_cor,
            slice_start=norm_anat_inf_slice_start_cor,
            slice_step=norm_anat_slices_gap_cor,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=50 * mm)
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 6 - Recall task############
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>Mémoire rappel </b> - "
                "Carte Statistique <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=10 >"
                "<b>Activation cérébrales pour la tâche de rappel libre</b>"
                "(reconnaisance des mots entendus lors de la première tâche "
                "à partir du son et générartion de la même phrase):"
                "</font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_recall,
            fig_rows=norm_anat_fig_rows,
            fig_cols=norm_anat_fig_cols,
            slice_start=norm_anat_inf_slice_start,
            slice_step=norm_anat_slices_gap,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=127 * mm)
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            data_2=self.spmT_recall,
            plane="sag",
            fig_rows=norm_anat_fig_rows_sag,
            fig_cols=norm_anat_fig_cols_sag,
            slice_start=norm_anat_inf_slice_start_sag,
            slice_step=norm_anat_slices_gap_sag,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=95 * mm)
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 7 - Explications protocole
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>Protocole GE2REC </b> - "
                "Informations <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))

        self.report.append(
            Paragraph(
                "<font size=12 > Le protocole GE2REC est composé de trois "
                "tâches interdépendantes : la génération de phrases avec "
                "encodage implicite (GE) et deux tâches de mémoire "
                "de rappel (2REC) qui sont la reconnaissance (RECO) "
                "et le rappel (RE) explication <br/> </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 > Ce protocole permet de cartographier "
                "l’interaction des fonctions du langage et de la mémoire "
                "(LMN language memory network) au niveau individuel."
                "Il fournit une évaluation exhaustive en incluant "
                "des modalités verbales, visuelles et divers processus "
                "langagiers et mémoriels. </font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 > <b> Première tâche – Stimuli et tâche "
                "de génération de phrases avec encodage implicite "
                "(GE)</b><br/>"
                "Il s'agit d'une séquence en bloc de génération de "
                "phrases avec encodage implicite (durée de 7,3 minutes)."
                "Les sujets écoutent des mots à travers un casque et doivent "
                "générer des phrases de manière implicite. Lors des périodes "
                "de contrôle, des pseudo-mots sont diffusés au sujet afin "
                "qu’il n’y ait pas de génération. "
                "Une croix de fixation apparaît lors des périodes de repos."
                "</font> ",
                self.styles["Left"],
            )
        )

        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 > <b> Deuxième tâche – Reconnaissance (RECO)"
                "</b><br/>"
                "La deuxième tâche est un paradigme évènementiel "
                "(durée de 6,8min). "
                "Des images sont présentées aux sujets selon un mode "
                "pseudo-aléatoire de 2,5 secondes. "
                "Ils doivent alors indiquer s’ils reconnaissent les images "
                "des objets dont les noms ont été diffusés lors de "
                "la première tâche GE. Dans le cadre clinique,"
                "le choix proposé est binaire. "
                "Soit le sujet indique qu’il reconnaît l’image (OLD), "
                "soit qu’il s’agit d’un nouvel item (NEW)."
                "</font> ",
                self.styles["Left"],
            )
        )

        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 > <b> Troisième tâche – Rappel (RA)"
                "</b><br/>"
                "La troisième tâche est un paradigme en bloc en modalité "
                "auditive (durée de 4,17 minutes). Les participants "
                "entendent les mots de la tâche GE et doivent se souvenir "
                "de manière explicite des phrases précédemment "
                "générées lors de la première tâche."
                "</font> ",
                self.styles["Left"],
            )
        )

        self.report.append(PageBreak())

        # page 8 - Anat -QC
        #################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRM anatomique </b> - "
                "Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Donnée d'entrée:</b></font> "
                f"<font size=10><i>{self.norm_anat}</i></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Paramètres d'acquistions:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom du procotocole d'acquistion: </b> "
                f"</font> {self.dict4runtime['norm_anat']['ProtocolName']}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom de la séquence: </b> "
                f"</font> {self.dict4runtime['norm_anat']['SequenceName']}",
                self.styles["Bullet1"],
            )
        )

        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_anat"]["SliceThickness"]

        if not sl_thick == "Undefined":
            sl_thick = sl_thick[0]

        st_end_sl = self.dict4runtime["norm_anat"]["Start/end slice"]

        if not st_end_sl == "Undefined" and st_end_sl == [0, 0]:
            st_end_sl = 0.0

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Epaisseure de coupe "
                f"/ écart entre les coupes [mm]:</b> "
                f"</font> {sl_thick} / {st_end_sl}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_anat"]["FOV"]

        if fov == "Undefined":
            fov = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> FOV (ap / fh / rl) [mm]:</b> </font> "
                f"{fov[0]} / {fov[1]} / {fov[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(Spacer(0 * mm, 1 * mm))
        vox_size = self.dict4runtime["norm_anat"][
            "Grid spacings (X,Y,Z,T,...)"
        ]
        if vox_size == "Undefined":
            vox_size = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in vox_size):
            vox_size = [round(elt, 1) for elt in vox_size]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:"
                f"</b> </font> {vox_size[0]} / {vox_size[1]} / {vox_size[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_anat"]["RepetitionTime"]
        te = self.dict4runtime["norm_anat"]["EchoTime"]
        flipang = self.dict4runtime["norm_anat"]["FlipAngle"]

        if flipang != "Undefined":
            flipang = round(flipang[0], 1)

        if te != "Undefined":
            te = round(te[0], 1)

        if tr != "Undefined":
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / Image flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            fig_rows=norm_anat_fig_rows,
            fig_cols=norm_anat_fig_cols,
            slice_start=norm_anat_inf_slice_start,
            slice_step=norm_anat_slices_gap,
            cmap_1=norm_anat_cmap,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=127 * mm)
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            plane="sag",
            fig_rows=norm_anat_fig_rows_sag,
            fig_cols=norm_anat_fig_cols_sag,
            slice_start=norm_anat_inf_slice_start_sag,
            slice_step=norm_anat_slices_gap_sag,
            cmap_1=norm_anat_cmap,
            cmap_2=spmT_cmap,
            vmin_2=self.spmT_vmin,
            vmax_2=self.spmT_vmax,
            out_dir=tmpdir.name,
        )
        slices_image = Image(slices_image, width=177 * mm, height=80 * mm)
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # page 9 - fMRI gene -QC
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de génération </b> "
                "- Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Donnée d'entrée:</b></font> "
                f"<font size=10><i>{self.norm_func_gene}</i></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Paramètres d'acquistions:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom du procotocole d'acquistion: </b> "
                f"</font> "
                f"{self.dict4runtime['norm_func_gene']['ProtocolName']}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom de la séquence: </b> "
                f"</font> "
                f"{self.dict4runtime['norm_func_gene']['SequenceName']}",
                self.styles["Bullet1"],
            )
        )

        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_func_gene"]["SliceThickness"]

        if not sl_thick == "Undefined":
            sl_thick = sl_thick[0]

        st_end_sl = self.dict4runtime["norm_func_gene"]["Start/end slice"]

        if not st_end_sl == "Undefined" and st_end_sl == [0, 0]:
            st_end_sl = 0.0

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Epaisseure de coupe "
                f"/ écart entre les coupes [mm]:</b> "
                f"</font> {sl_thick} / {st_end_sl}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_func_gene"]["FOV"]

        if fov == "Undefined":
            fov = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> FOV (ap / fh / rl) [mm]:</b> </font> "
                f"{fov[0]} / {fov[1]} / {fov[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(Spacer(0 * mm, 1 * mm))
        vox_size = self.dict4runtime["norm_func_gene"][
            "Grid spacings (X,Y,Z,T,...)"
        ]
        if vox_size == "Undefined":
            vox_size = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in vox_size):
            vox_size = [round(elt, 1) for elt in vox_size]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:"
                f"</b> </font> {vox_size[0]} / {vox_size[1]} / {vox_size[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_func_gene"]["RepetitionTime"]
        te = self.dict4runtime["norm_func_gene"]["EchoTime"]
        flipang = self.dict4runtime["norm_func_gene"]["FlipAngle"]

        if flipang != "Undefined":
            flipang = round(flipang[0], 1)

        if te != "Undefined":
            te = round(te[0], 1)

        if tr != "Undefined":
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / Image flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b> Image fonctionnelle normalisée (MNI) "
                "(1er dynamique):<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_func_gene,
            fig_rows=norm_func_fig_rows,
            fig_cols=norm_func_fig_cols,
            slice_start=norm_func_inf_slice_start,
            slice_step=norm_func_slices_gap,
            cmap_1=norm_func_cmap,
            out_dir=tmpdir.name,
        )

        slices_image = Image(
            slices_image, width=7.4803 * inch, height=7.2 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 10 - fMRI gene -Qc
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de génération "
                "</b> - Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))

        # Carpet plot
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Intensité des voxels au cours du "
                "temps <i>(image normalisée)</i>:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        carpet = plot_carpet(
            self.norm_func_gene,
            self.norm_func_mask,
            t_r=tr / 1000,
            standardize="zscore_sample",
            title="global patterns over time",
        )
        out_carpet_plot = os.path.join(
            tmpdir.name,
            "gene_carpet.png",
        )
        carpet.savefig(out_carpet_plot, format="png", dpi=200)
        im_carpet = Image(out_carpet_plot, 6.468 * inch, 3.018 * inch)
        im_carpet.hAlign = "CENTER"
        self.report.append(im_carpet)

        # Realignment
        sources_images_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "sources_images"
        )
        im_check = None
        im_tra = None
        im_rot = None
        out_file_tra = os.path.join(
            tmpdir.name,
            "gene_QC_translation.png",
        )
        out_file_rot = os.path.join(
            tmpdir.name,
            "gene_QC_rotation.png",
        )
        qc = plot_realignment_parameters(
            self.realignment_parameters_gene,
            vox_size,
            out_file_tra,
            out_file_rot,
        )

        if qc == 0:
            # 912px × 892px
            im_check = Image(
                os.path.join(sources_images_dir, "No-check-mark.png"),
                13.0 * mm,
                12.7 * mm,
            )
        elif qc == 1:
            # 940px × 893px
            im_check = Image(
                os.path.join(sources_images_dir, "OK-check-mark.png"),
                13.0 * mm,
                12.4 * mm,
            )

        im_tra = Image(out_file_tra, 6.468 * inch, 3.018 * inch)
        im_rot = Image(out_file_rot, 6.468 * inch, 3.018 * inch)
        message = Paragraph(
            "<font size=14 > <b> Mouvements </b> </font>",
            self.styles["Left"],
        )
        im_check.hAlign = "CENTER"
        title = [[message, im_check]]
        t = Table(title, [110 * mm, 30 * mm])  # colWidths, rowHeight
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        t.hAlign = "LEFT"
        self.report.append(t)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_tra.hAlign = "CENTER"
        self.report.append(im_tra)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_rot.hAlign = "CENTER"
        self.report.append(im_rot)
        self.report.append(PageBreak())

        # page 11 - fMRI reco -QC
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de reconnaissance </b> "
                "- Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Donnée d'entrée:</b></font> "
                f"<font size=10><i>{self.norm_func_reco}</i></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Paramètres d'acquistions:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom du procotocole d'acquistion: </b> "
                f"</font>"
                f"{self.dict4runtime['norm_func_reco']['ProtocolName']}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom de la séquence: </b> </font>"
                f"{self.dict4runtime['norm_func_reco']['SequenceName']}",
                self.styles["Bullet1"],
            )
        )

        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_func_reco"]["SliceThickness"]

        if not sl_thick == "Undefined":
            sl_thick = sl_thick[0]

        st_end_sl = self.dict4runtime["norm_func_reco"]["Start/end slice"]

        if not st_end_sl == "Undefined" and st_end_sl == [0, 0]:
            st_end_sl = 0.0

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Epaisseure de coupe "
                f"/ écart entre les coupes [mm]:</b> "
                f"</font> {sl_thick} / {st_end_sl}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_func_reco"]["FOV"]

        if fov == "Undefined":
            fov = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> FOV (ap / fh / rl) [mm]:</b> </font> "
                f"{fov[0]} / {fov[1]} / {fov[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(Spacer(0 * mm, 1 * mm))
        vox_size = self.dict4runtime["norm_func_reco"][
            "Grid spacings (X,Y,Z,T,...)"
        ]
        if vox_size == "Undefined":
            vox_size = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in vox_size):
            vox_size = [round(elt, 1) for elt in vox_size]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:"
                f"</b> </font> {vox_size[0]} / {vox_size[1]} / {vox_size[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_func_reco"]["RepetitionTime"]
        te = self.dict4runtime["norm_func_reco"]["EchoTime"]
        flipang = self.dict4runtime["norm_func_reco"]["FlipAngle"]

        if flipang != "Undefined":
            flipang = round(flipang[0], 1)

        if te != "Undefined":
            te = round(te[0], 1)

        if tr != "Undefined":
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / Image flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b> Image fonctionnelle normalisée (MNI) "
                "(1er dynamique):<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_func_reco,
            fig_rows=norm_func_fig_rows,
            fig_cols=norm_func_fig_cols,
            slice_start=norm_func_inf_slice_start,
            slice_step=norm_func_slices_gap,
            cmap_1=norm_func_cmap,
            out_dir=tmpdir.name,
        )

        slices_image = Image(
            slices_image, width=7.4803 * inch, height=7.2 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 12 - fMRI reco -Qc
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de reconnaissance "
                "</b> - Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))

        # Carpet plot
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Intensité des voxels au cours du "
                "temps <i>(image normalisée)</i>:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        carpet = plot_carpet(
            self.norm_func_reco,
            self.norm_func_mask,
            t_r=tr / 1000,
            standardize="zscore_sample",
            title="global patterns over time",
        )
        out_carpet_plot = os.path.join(
            tmpdir.name,
            "reco_carpet.png",
        )
        carpet.savefig(out_carpet_plot, format="png", dpi=200)
        im_carpet = Image(out_carpet_plot, 6.468 * inch, 3.018 * inch)
        im_carpet.hAlign = "CENTER"
        self.report.append(im_carpet)

        # Realignment
        sources_images_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "sources_images"
        )
        im_check = None
        im_tra = None
        im_rot = None
        out_file_tra = os.path.join(
            tmpdir.name,
            "reco_QC_translation.png",
        )
        out_file_rot = os.path.join(
            tmpdir.name,
            "reco_QC_rotation.png",
        )
        qc = plot_realignment_parameters(
            self.realignment_parameters_reco,
            vox_size,
            out_file_tra,
            out_file_rot,
        )

        if qc == 0:
            # 912px × 892px
            im_check = Image(
                os.path.join(sources_images_dir, "No-check-mark.png"),
                13.0 * mm,
                12.7 * mm,
            )
        elif qc == 1:
            # 940px × 893px
            im_check = Image(
                os.path.join(sources_images_dir, "OK-check-mark.png"),
                13.0 * mm,
                12.4 * mm,
            )

        im_tra = Image(out_file_tra, 6.468 * inch, 3.018 * inch)
        im_rot = Image(out_file_rot, 6.468 * inch, 3.018 * inch)
        message = Paragraph(
            "<font size=14 > <b> Mouvements </b> </font>",
            self.styles["Left"],
        )
        im_check.hAlign = "CENTER"
        title = [[message, im_check]]
        t = Table(title, [110 * mm, 30 * mm])  # colWidths, rowHeight
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        t.hAlign = "LEFT"
        self.report.append(t)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_tra.hAlign = "CENTER"
        self.report.append(im_tra)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_rot.hAlign = "CENTER"
        self.report.append(im_rot)
        self.report.append(PageBreak())

        # page 13- fMRI recall -QC
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de rappel </b> "
                "- Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Donnée d'entrée:</b></font> "
                f"<font size=10><i>{self.norm_func_recall}</i></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Paramètres d'acquistions:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom du procotocole d'acquistion: </b> "
                f"</font> "
                f"{self.dict4runtime['norm_func_recall']['ProtocolName']}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                f"<font size=11> <b> Nom de la séquence: </b> "
                f"</font> "
                f"{self.dict4runtime['norm_func_recall']['SequenceName']}",
                self.styles["Bullet1"],
            )
        )

        self.report.append(Spacer(0 * mm, 1 * mm))
        sl_thick = self.dict4runtime["norm_func_recall"]["SliceThickness"]

        if not sl_thick == "Undefined":
            sl_thick = sl_thick[0]

        st_end_sl = self.dict4runtime["norm_func_recall"]["Start/end slice"]

        if not st_end_sl == "Undefined" and st_end_sl == [0, 0]:
            st_end_sl = 0.0

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Epaisseure de coupe "
                f"/ écart entre les coupes [mm]:</b> "
                f"</font> {sl_thick} / {st_end_sl}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        fov = self.dict4runtime["norm_func_recall"]["FOV"]

        if fov == "Undefined":
            fov = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in fov):
            fov = [round(elt, 1) for elt in fov]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> FOV (ap / fh / rl) [mm]:</b> </font> "
                f"{fov[0]} / {fov[1]} / {fov[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(Spacer(0 * mm, 1 * mm))
        vox_size = self.dict4runtime["norm_func_recall"][
            "Grid spacings (X,Y,Z,T,...)"
        ]
        if vox_size == "Undefined":
            vox_size = ["Undefined"] * 3

        if all(isinstance(elt, (int, float)) for elt in vox_size):
            vox_size = [round(elt, 1) for elt in vox_size]

        self.report.append(
            Paragraph(
                f"<font size=11> <b> Voxel size (x / y / z) [mm]:"
                f"</b> </font> {vox_size[0]} / {vox_size[1]} / {vox_size[2]}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tr = self.dict4runtime["norm_func_recall"]["RepetitionTime"]
        te = self.dict4runtime["norm_func_recall"]["EchoTime"]
        flipang = self.dict4runtime["norm_func_recall"]["FlipAngle"]

        if flipang != "Undefined":
            flipang = round(flipang[0], 1)

        if te != "Undefined":
            te = round(te[0], 1)

        if tr != "Undefined":
            tr = round(tr[0], 1)

        self.report.append(
            Paragraph(
                f"<font size=11> <b> TR [ms] / TE [ms] / Image flip angle"
                f" [deg]:</b> </font> {tr} / {te} / {flipang}",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(
                "<font size=12 ><b> Image fonctionnelle normalisée (MNI) "
                "(1er dynamique):<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(
            Paragraph(
                self.neuro_conv,
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_func_recall,
            fig_rows=norm_func_fig_rows,
            fig_cols=norm_func_fig_cols,
            slice_start=norm_func_inf_slice_start,
            slice_step=norm_func_slices_gap,
            cmap_1=norm_func_cmap,
            out_dir=tmpdir.name,
        )

        slices_image = Image(
            slices_image, width=7.4803 * inch, height=7.2 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)

        self.report.append(PageBreak())

        # page 14 - fMRI recall -Qc
        ######################################################################
        self.report.append(
            Paragraph(
                "<font size=18 ><b>IRMf tâche de rappel "
                "</b> - Contrôle qualité <br/></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))

        # Carpet plot
        self.report.append(
            Paragraph(
                "<font size=12 ><b>Intensité des voxels au cours du "
                "temps <i>(image normalisée)</i>:<br/></b></font>",
                self.styles["Left"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        carpet = plot_carpet(
            self.norm_func_recall,
            self.norm_func_mask,
            t_r=tr / 1000,
            standardize="zscore_sample",
            title="global patterns over time",
        )
        out_carpet_plot = os.path.join(
            tmpdir.name,
            "recall_carpet.png",
        )
        carpet.savefig(out_carpet_plot, format="png", dpi=200)
        im_carpet = Image(out_carpet_plot, 6.468 * inch, 3.018 * inch)
        im_carpet.hAlign = "CENTER"
        self.report.append(im_carpet)

        # Realignment
        sources_images_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "sources_images"
        )
        im_check = None
        im_tra = None
        im_rot = None
        out_file_tra = os.path.join(
            tmpdir.name,
            "recall_QC_translation.png",
        )
        out_file_rot = os.path.join(
            tmpdir.name,
            "recall_QC_rotation.png",
        )
        qc = plot_realignment_parameters(
            self.realignment_parameters_recall,
            vox_size,
            out_file_tra,
            out_file_rot,
        )

        if qc == 0:
            # 912px × 892px
            im_check = Image(
                os.path.join(sources_images_dir, "No-check-mark.png"),
                13.0 * mm,
                12.7 * mm,
            )
        elif qc == 1:
            # 940px × 893px
            im_check = Image(
                os.path.join(sources_images_dir, "OK-check-mark.png"),
                13.0 * mm,
                12.4 * mm,
            )

        im_tra = Image(out_file_tra, 6.468 * inch, 3.018 * inch)
        im_rot = Image(out_file_rot, 6.468 * inch, 3.018 * inch)
        message = Paragraph(
            "<font size=14 > <b> Mouvements " " </b> </font>",
            self.styles["Left"],
        )
        im_check.hAlign = "CENTER"
        title = [[message, im_check]]
        t = Table(title, [110 * mm, 30 * mm])  # colWidths, rowHeight
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        t.hAlign = "LEFT"
        self.report.append(t)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_tra.hAlign = "CENTER"
        self.report.append(im_tra)
        self.report.append(Spacer(0 * mm, 4 * mm))
        im_rot.hAlign = "CENTER"
        self.report.append(im_rot)
        self.report.append(PageBreak())

        self.page.build(self.report, canvasmaker=PageNumCanvas)
        tmpdir.cleanup()

    def mriqc_anat_make_report(
        self,
    ):
        """Make mriqc anat individual report"""

        # First page - cover ##################################################
        #######################################################################

        self.report.append(self.image_cov)
        # width, height
        self.report.append(Spacer(0 * mm, 8 * mm))
        self.report.append(Paragraph(self.header_title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(Paragraph(self.title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(self.cover_data)
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(self.textDisclaimer, self.styles["Justify"])
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        self.report.append(
            Paragraph(
                "<font size = 8><sup>$</sup>Esteban O et "
                "al., <i>MRIQC: Advancing the Automatic "
                "Prediction of Image Quality in MRI from"
                " Unseen Sites</i>, PLOS ONE 12(9)"
                ":e0184661.</font>",
                self.styles["Left"],
            )
        )
        self.report.append(PageBreak())

        # Second page - IQMs ##################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Spatial resolution ################################################
        self.report.append(
            Paragraph(
                "<font size = 15 > <b>SPATIAL RESOLUTION</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(Paragraph("Length - Spacing", self.styles["Left2"]))
        self.report.append(Spacer(0 * mm, 10 * mm))

        # size_x
        self.report.append(self.get_iqms_data("size_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # size_y
        self.report.append(self.get_iqms_data("size_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # size_z
        self.report.append(self.get_iqms_data("size_z"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # spacing_x
        self.report.append(self.get_iqms_data("spacing_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # spacing_y
        self.report.append(self.get_iqms_data("spacing_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # spacing_z
        self.report.append(self.get_iqms_data("spacing_z"))
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Signal-to-noise ###################################################
        self.report.append(
            Paragraph(
                "<font size = 15 > <b>NOISE</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "Impact of noise and/or evaluation of "
                "the fitness of a noise model",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # snr_csf
        self.report.append(self.get_iqms_data("snr_csf"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snr_wm
        self.report.append(self.get_iqms_data("snr_wm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snr_gm
        self.report.append(self.get_iqms_data("snr_gm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snr_total
        self.report.append(self.get_iqms_data("snr_total"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # snrd_csf
        self.report.append(self.get_iqms_data("snrd_csf"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snrd_wm
        self.report.append(self.get_iqms_data("snrd_wm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snrd_gm
        self.report.append(self.get_iqms_data("snrd_gm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # snrd_total
        self.report.append(self.get_iqms_data("snrd_total"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # cnr
        self.report.append(self.get_iqms_data("cnr"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # qi_2
        self.report.append(self.get_iqms_data("qi_2"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # cjv
        self.report.append(self.get_iqms_data("cjv"))
        self.report.append(Spacer(0 * mm, 52 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("snrd_csf", "cnr", "qi_2", "cjv"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # Third page - IQMs ###################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Spatial distribution ##############################################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>SPATIAL DISTRIBUTION </b></font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "Information theory to evaluate the "
                "spatial distribution of information",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # fber
        self.report.append(self.get_iqms_data("fber"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # efc
        self.report.append(self.get_iqms_data("efc"))
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Artifacts #########################################################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>ARTIFACTS </b></font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "Estimates artefacts and signal leakage due to "
                "rapid movement (e.g. eye movement or blood "
                "vessel pulsation, etc.)",
                self.styles["Left2"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        # wm2max
        self.report.append(self.get_iqms_data("wm2max"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # qi_1
        self.report.append(self.get_iqms_data("qi_1"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # inu_range
        self.report.append(self.get_iqms_data("inu_range"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # inu_med
        self.report.append(self.get_iqms_data("inu_med"))
        self.report.append(Spacer(0 * mm, 100 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("fber", "efc", "wm2max", "qi_1", "inu_range"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # fourth page - IQMs ##################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Tissues Quality ###################################################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>TISSUES QUALITY </b></font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "Metrics that do not fall into the above "
                "categories: statistical properties of "
                "tissue distributions, volume overlap "
                "of tissues, image harpness/blurriness,"
                " etc.",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # summary_csf_mean
        self.report.append(self.get_iqms_data("summary_csf_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_stdv
        self.report.append(self.get_iqms_data("summary_csf_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_median
        self.report.append(self.get_iqms_data("summary_csf_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_mad
        self.report.append(self.get_iqms_data("summary_csf_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_p95
        self.report.append(self.get_iqms_data("summary_csf_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_p05
        self.report.append(self.get_iqms_data("summary_csf_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_k
        self.report.append(self.get_iqms_data("summary_csf_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_csf_n
        self.report.append(self.get_iqms_data("summary_csf_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # summary_gm_mean
        self.report.append(self.get_iqms_data("summary_gm_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_stdv
        self.report.append(self.get_iqms_data("summary_gm_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_median
        self.report.append(self.get_iqms_data("summary_gm_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_mad
        self.report.append(self.get_iqms_data("summary_gm_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_p95
        self.report.append(self.get_iqms_data("summary_gm_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_p05
        self.report.append(self.get_iqms_data("summary_gm_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_k
        self.report.append(self.get_iqms_data("summary_gm_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_gm_n
        self.report.append(self.get_iqms_data("summary_gm_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # summary_wm_mean
        self.report.append(self.get_iqms_data("summary_wm_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_stdv
        self.report.append(self.get_iqms_data("summary_wm_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_median
        self.report.append(self.get_iqms_data("summary_wm_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_mad
        self.report.append(self.get_iqms_data("summary_wm_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_p95
        self.report.append(self.get_iqms_data("summary_wm_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_p05
        self.report.append(self.get_iqms_data("summary_wm_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_k
        self.report.append(self.get_iqms_data("summary_wm_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_wm_n
        self.report.append(self.get_iqms_data("summary_wm_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # summary_bg_mean
        self.report.append(self.get_iqms_data("summary_bg_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_stdv
        self.report.append(self.get_iqms_data("summary_bg_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_median
        self.report.append(self.get_iqms_data("summary_bg_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_mad
        self.report.append(self.get_iqms_data("summary_bg_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_p95
        self.report.append(self.get_iqms_data("summary_bg_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_p05
        self.report.append(self.get_iqms_data("summary_bg_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_k
        self.report.append(self.get_iqms_data("summary_bg_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_n
        self.report.append(self.get_iqms_data("summary_bg_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # fwhm_x
        self.report.append(self.get_iqms_data("fwhm_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_y
        self.report.append(self.get_iqms_data("fwhm_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_z
        self.report.append(self.get_iqms_data("fwhm_z"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_avg
        self.report.append(self.get_iqms_data("fwhm_avg"))
        self.report.append(Spacer(0 * mm, 7 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("summary_csf_k", "fwhm_x"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # Fifth page page - IQMs###############################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Tissues Quality ###################################################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>TISSUES QUALITY</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(
            Paragraph(
                "Metrics that do not fall into the above categories: "
                "statistical properties of tissue distributions, "
                "volume overlap of tissues, image harpness/blurriness, "
                "etc.",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # icvs_csf
        self.report.append(self.get_iqms_data("icvs_csf"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # icvs_gm
        self.report.append(self.get_iqms_data("icvs_gm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # icvs_wm
        self.report.append(self.get_iqms_data("icvs_wm"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # rpve_csf
        self.report.append(self.get_iqms_data("rpve_csf"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # rpve_gm
        self.report.append(self.get_iqms_data("rpve_gm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # rpve_wm
        self.report.append(self.get_iqms_data("rpve_wm"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # tpm_overlap_csf
        self.report.append(self.get_iqms_data("tpm_overlap_csf"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # tpm_overlap_gm
        self.report.append(self.get_iqms_data("tpm_overlap_gm"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # tpm_overlap_wm
        self.report.append(self.get_iqms_data("tpm_overlap_wm"))
        self.report.append(Spacer(0 * mm, 147 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("rpve_csf", "tpm_overlap_csf"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # Sixth page - slice planes display - Raw anatomic ####################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 >Raw anatomic images</font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 ><i>"Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain.</i></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tmpdir = tempfile.TemporaryDirectory()
        slices_image = plot_slice_planes(
            data_1=self.anat,
            fig_rows=self.anat_fig_rows,
            fig_cols=self.anat_fig_cols,
            slice_start=self.anat_inf_slice_start,
            slice_step=self.anat_slices_gap,
            cmap_1="Greys_r",
            out_dir=tmpdir.name,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Seventh page - slice planes display - Normalised anatomic (MNI) #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 >Normalised anatomic (MNI)</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_anat,
            fig_rows=self.norm_anat_fig_rows,
            fig_cols=self.norm_anat_fig_cols,
            slice_start=self.norm_anat_inf_slice_start,
            slice_step=self.norm_anat_slices_gap,
            cmap_1="Greys_r",
            out_dir=tmpdir.name,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Eighth page - slice planes display - Background #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 > Anatomical "
                "image with background enhancement</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.anat,
            fig_rows=self.anat_fig_rows,
            fig_cols=self.anat_fig_cols,
            slice_start=self.anat_inf_slice_start,
            slice_step=self.anat_slices_gap,
            cmap_1="viridis_r",
            out_dir=tmpdir.name,
            only_noise=True,
            out_name="background",
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Nineth page - slice planes display - Segmentation #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 > Anatomical "
                "image with segmentation</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Brain tissue segmentation",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        plot_seg = plot_segmentation(
            self.anat,
            self.segmentation,
            out_dir=tmpdir.name,
            name="PlotSegmentation",
            cut_coords=9,
            display_mode="z",
            levels=[0.5, 1.5, 2.5],
            colors=["r", "g", "b"],
        )

        image = Image(plot_seg, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Brain mask",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        plot_bmask = plot_segmentation(
            self.anat,
            self.brain_mask,
            out_dir=tmpdir.name,
            name="PlotBrainmask",
            cut_coords=9,
            display_mode="z",
            levels=[0.5],
            colors=["r"],
        )

        image = Image(plot_bmask, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Hat mask",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        plot_airmask = plot_segmentation(
            self.anat,
            self.air_mask,
            out_dir=tmpdir.name,
            name="PlotAirmask",
            cut_coords=6,
            display_mode="x",
            levels=[0.5],
            colors=["r"],
        )

        image = Image(plot_airmask, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Head outline",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        plot_headmask = plot_segmentation(
            self.anat,
            self.head_mask,
            out_dir=tmpdir.name,
            name="PlotHeadmask",
            cut_coords=6,
            display_mode="x",
            levels=[0.5],
            colors=["r"],
        )

        image = Image(plot_headmask, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Artifacts in background",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        plot_artmask = plot_segmentation(
            self.anat,
            self.art_mask,
            out_dir=tmpdir.name,
            name="PlotArtmask",
            cut_coords=9,
            display_mode="z",
            levels=[0.5],
            colors=["r"],
            saturate=True,
        )

        image = Image(plot_artmask, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.report.append(PageBreak())

        # Tenth page - qi2 plot #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>Distribution of the noise in "
                "the background</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))

        qi2 = plot_qi2(
            np.asarray(self.iqms_data["histogram_qi2_x_grid"]),
            np.asarray(self.iqms_data["histogram_qi2_ref_pdf"]),
            np.asarray(self.iqms_data["histogram_qi2_fit_pdf"]),
            np.asarray(self.iqms_data["histogram_qi2_ref_data"]),
            int(self.iqms_data["histogram_qi2_cutoff_idx"]),
            out_file=os.path.join(
                tmpdir.name,
                "qi2_plot.png",
            ),
        )
        image = Image(qi2, width=7.4803 * inch, height=5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.page.build(self.report, canvasmaker=PageNumCanvas)
        tmpdir.cleanup()

    def mriqc_func_make_report(
        self,
    ):
        """Make mriqc functional individual report"""

        # First page - cover ##################################################
        #######################################################################

        self.report.append(self.image_cov)
        # width, height
        self.report.append(Spacer(0 * mm, 8 * mm))
        self.report.append(Paragraph(self.header_title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(Paragraph(self.title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(self.cover_data)
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(self.textDisclaimer, self.styles["Justify"])
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        self.report.append(
            Paragraph(
                "<font size = 8><sup>$</sup>Esteban O et "
                "al., <i>MRIQC: Advancing the Automatic "
                "Prediction of Image Quality in MRI from"
                " Unseen Sites</i>, PLOS ONE 12(9)"
                ":e0184661.</font>",
                self.styles["Left"],
            )
        )
        self.report.append(PageBreak())

        # Second page - IQMs ##################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Spatial and temporal resolution ###################################
        self.report.append(
            Paragraph(
                "<font size = 15 > <b>SPATIAL AND "
                "TEMPORAL RESOLUTION</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 2 * mm))
        self.report.append(Paragraph("Length - Spacing", self.styles["Left2"]))
        self.report.append(Spacer(0 * mm, 10 * mm))

        # size_x
        self.report.append(self.get_iqms_data("size_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # size_y
        self.report.append(self.get_iqms_data("size_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # size_z
        self.report.append(self.get_iqms_data("size_z"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # size_t
        self.report.append(self.get_iqms_data("size_t"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # spacing_x
        self.report.append(self.get_iqms_data("spacing_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # spacing_y
        self.report.append(self.get_iqms_data("spacing_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # spacing_z
        self.report.append(self.get_iqms_data("spacing_z"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # spacing_tr
        self.report.append(self.get_iqms_data("spacing_tr"))
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Measures for artifacts and other ##################################
        self.report.append(
            Paragraph(
                "<font size = 15 > <b>MEASURES FOR "
                "ARTIFACTS AND OTHER</b> </font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # aor
        self.report.append(self.get_iqms_data("aor"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # aqi
        self.report.append(self.get_iqms_data("aqi"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # gsr_x
        self.report.append(self.get_iqms_data("gsr_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # gsr_y
        self.report.append(self.get_iqms_data("gsr_y"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # fd_mean
        self.report.append(self.get_iqms_data("fd_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fd_num
        self.report.append(self.get_iqms_data("fd_num"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fd_perc
        self.report.append(self.get_iqms_data("fd_perc"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # dummy_trs
        self.report.append(self.get_iqms_data("dummy_trs"))
        self.report.append(Spacer(0 * mm, 82 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("gsr_x", "fd_mean"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # Third page - IQMs ###################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Measures for the spatial information ##############################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>MEASURES FOR THE "
                "SPATIAL INFORMATION</b></font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # summary_bg_mean
        self.report.append(self.get_iqms_data("summary_bg_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_stdv
        self.report.append(self.get_iqms_data("summary_bg_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_median
        self.report.append(self.get_iqms_data("summary_bg_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_mad
        self.report.append(self.get_iqms_data("summary_bg_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_p95
        self.report.append(self.get_iqms_data("summary_bg_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_p05
        self.report.append(self.get_iqms_data("summary_bg_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_k
        self.report.append(self.get_iqms_data("summary_bg_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_bg_n
        self.report.append(self.get_iqms_data("summary_bg_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # summary_fg_mean
        self.report.append(self.get_iqms_data("summary_fg_mean"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_stdv
        self.report.append(self.get_iqms_data("summary_fg_stdv"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_median
        self.report.append(self.get_iqms_data("summary_fg_median"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_mad
        self.report.append(self.get_iqms_data("summary_fg_mad"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_p95
        self.report.append(self.get_iqms_data("summary_fg_p95"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_p05
        self.report.append(self.get_iqms_data("summary_fg_p05"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_k
        self.report.append(self.get_iqms_data("summary_fg_k"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # summary_fg_n
        self.report.append(self.get_iqms_data("summary_fg_n"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # snr
        self.report.append(self.get_iqms_data("snr"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # fber
        self.report.append(self.get_iqms_data("fber"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # efc
        self.report.append(self.get_iqms_data("efc"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # fwhm_x
        self.report.append(self.get_iqms_data("fwhm_x"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_y
        self.report.append(self.get_iqms_data("fwhm_y"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_z
        self.report.append(self.get_iqms_data("fwhm_z"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # fwhm_avg
        self.report.append(self.get_iqms_data("fwhm_avg"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        self.report.append(Spacer(0 * mm, 66 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("summary_bg_k", "fber", "efc", "fwhm_x"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # fourth page - IQMs ##################################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 ><b>Image parameters </b></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 20 * mm))
        # Measures for the temporal information #############################
        self.report.append(
            Paragraph(
                "<font size = 15 ><b>MEASURES FOR THE "
                "TEMPORAL INFORMATION</b></font>",
                self.styles["Bullet1"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        # dvars_nstd
        self.report.append(self.get_iqms_data("dvars_nstd"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # dvars_std
        self.report.append(self.get_iqms_data("dvars_std"))
        self.report.append(Spacer(0 * mm, 1 * mm))
        # dvars_vstd
        self.report.append(self.get_iqms_data("dvars_vstd"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # tsnr
        self.report.append(self.get_iqms_data("tsnr"))
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        # gcor
        self.report.append(self.get_iqms_data("gcor"))
        self.report.append(Spacer(0 * mm, 170 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))

        for param in ("dvars_nstd", "tsnr", "gcor"):
            self.report.append(
                Paragraph(
                    "<font size = 8><sup>"
                    + self.dict4runtime["extra_info"][param][2]
                    + "</sup>"
                    + self.dict4runtime["extra_info"][param][3]
                    + "</font>",
                    self.styles["Left"],
                )
            )

        self.report.append(PageBreak())

        # Fifth - Carpet plot #######################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>Summary plot </b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 9 > <i>Summary plot, "
                "showing the slice-wise signal intensity at the extremes for "
                "the identification of spikes, the outliers metric, "
                "the DVARS, the FD and the carpet plot. "
                "The carpet plot rows correspond to voxelwise time series, "
                "and are separated into regions: cortical gray matter, "
                "deep gray matter, white matter and cerebrospinal fluid, "
                "cerebellum and the brain-edge or “crown”. The crown "
                "corresponds to the voxels located on a closed "
                "band around the brain. Carpet plot done using the head "
                "motion corrected functional image. </i></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        # reminder: A4 == 210mmx297mm
        image = Image(
            self.IQMS_plot, width=7.4803 * inch, height=9.0551 * inch
        )
        image.hAlign = "CENTER"
        self.report.append(image)
        self.report.append(PageBreak())

        # Sixth - slice planes display - Raw functional #######################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 >Raw functional "
                "images (1<sup>st</sup> dynamic)"
                "</font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 ><i>"Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain.</i></font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        tmpdir = tempfile.TemporaryDirectory()
        slices_image = plot_slice_planes(
            data_1=self.func,
            fig_rows=self.func_fig_rows,
            fig_cols=self.func_fig_cols,
            slice_start=self.func_inf_slice_start,
            slice_step=self.func_slices_gap,
            cmap_1="Greys_r",
            out_dir=tmpdir.name,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Seventh page - slice planes display - Normalised functional (MNI) ###
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 >Normalised mean functional (MNI)</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.norm_func,
            fig_rows=self.norm_func_fig_rows,
            fig_cols=self.norm_func_fig_cols,
            slice_start=self.norm_func_inf_slice_start,
            slice_step=self.norm_func_slices_gap,
            cmap_1="Greys_r",
            out_dir=tmpdir.name,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Eighth page - slice planes display - stddev #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 > Temporal standard deviation</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.stddev,
            fig_rows=self.func_fig_rows,
            fig_cols=self.func_fig_cols,
            slice_start=self.func_inf_slice_start,
            slice_step=self.func_slices_gap,
            cmap_1="viridis",
            out_dir=tmpdir.name,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Ninth page - slice planes display - Background #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI axial slice "
                "planes display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 > Average functional "
                "image with background enhancement</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))
        slices_image = plot_slice_planes(
            data_1=self.func_mean,
            fig_rows=self.func_fig_rows,
            fig_cols=self.func_fig_cols,
            slice_start=self.func_inf_slice_start,
            slice_step=self.func_slices_gap,
            cmap_1="viridis_r",
            out_dir=tmpdir.name,
            only_noise=True,
        )
        # reminder: A4 == 210mmx297mm
        slices_image = Image(
            slices_image, width=7.4803 * inch, height=9.0551 * inch
        )
        slices_image.hAlign = "CENTER"
        self.report.append(slices_image)
        self.report.append(PageBreak())

        # Tenth page - slice planes display - Segmentation #####
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>MRI display</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "<font size = 14 > Mean functional "
                "image with segmentation</font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                '<font size = 9 > <i> "Neurological" '
                "convention, the left side of the "
                "image corresponds to the left side of "
                "the brain. </i> <br/> </font>",
                self.styles["Center"],
            )
        )

        self.report.append(Spacer(0 * mm, 5 * mm))
        self.report.append(
            Paragraph(
                "Brain mask",
                self.styles["Left2"],
            )
        )
        self.report.append(Spacer(0 * mm, 1 * mm))

        plot_bmask = plot_segmentation(
            self.func_mean,
            self.brain_mask,
            out_dir=tmpdir.name,
            name="PlotBrainmask",
            cut_coords=9,
            display_mode="z",
            levels=[0.5],
            colors=["r"],
        )

        image = Image(plot_bmask, width=7.4803 * inch, height=1.5 * inch)
        image.hAlign = "CENTER"
        self.report.append(image)

        self.page.build(self.report, canvasmaker=PageNumCanvas)
        tmpdir.cleanup()

    def mriqc_group_make_report(self):
        """Make mriqc group report"""

        # First page - cover ##################################################
        #######################################################################

        self.report.append(self.image_cov)
        # width, height
        self.report.append(Spacer(0 * mm, 8 * mm))
        self.report.append(Paragraph(self.header_title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(Paragraph(self.title, self.styles["Center"]))
        self.report.append(Spacer(0 * mm, 10 * mm))
        self.report.append(self.cover_data)
        self.report.append(Spacer(0 * mm, 6 * mm))
        self.report.append(
            Paragraph(self.textDisclaimer, self.styles["Justify"])
        )
        self.report.append(Spacer(0 * mm, 6 * mm))
        # Footnote
        line = ReportLine(500)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 2.5 * mm))
        self.report.append(
            Paragraph(
                "<font size = 8><sup>$</sup>Esteban O et "
                "al., <i>MRIQC: Advancing the Automatic "
                "Prediction of Image Quality in MRI from"
                " Unseen Sites</i>, PLOS ONE 12(9)"
                ":e0184661.</font>",
                self.styles["Left"],
            )
        )
        self.report.append(PageBreak())

        # Second page - Subject #########################################
        #######################################################################
        self.report.append(
            Paragraph(
                "<font size = 18 > <b>Subjects</b> </font>",
                self.styles["Center"],
            )
        )
        self.report.append(Spacer(0 * mm, 4 * mm))
        line = ReportLine(150)
        line.hAlign = "CENTER"
        self.report.append(line)
        self.report.append(Spacer(0 * mm, 4 * mm))

        data = [["Input data", "Site", "MRI scanner"]]
        for key in list(self.dict4runtime.keys()):
            data_file = []
            for i in ("FileName", "Site", "Spectro"):
                tag = self.dict4runtime[key][i]
                if len(tag) > 30:
                    tag = f"{tag[:30]}\n{tag[30:]}"
                data_file.append(tag)
            data.append(data_file)

        subjects_table = Table(data)
        subjects_table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, 0), "Times-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        self.report.append(subjects_table)
        self.report.append(PageBreak())

        modality = self.mriqc_group["modality"]
        if modality == "anat":
            # Third page - Boxplots Noise ####################################
            ##################################################################
            self.report.append(
                Paragraph(
                    "<font size = 18 > <b>Noise</b> </font>",
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))

            boxplot_snr_image = Image(
                self.mriqc_group["SNR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_snrd_image = Image(
                self.mriqc_group["SNRD"], width=3 * inch, height=2.5 * inch
            )
            boxplot_cnr_image = Image(
                self.mriqc_group["CNR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_cjv_image = Image(
                self.mriqc_group["CJV"], width=3 * inch, height=2.5 * inch
            )
            boxplot_qi2_image = Image(
                self.mriqc_group["QI"], width=3 * inch, height=2.5 * inch
            )

            noise_data = [
                [boxplot_snr_image, boxplot_snrd_image],
                [boxplot_cnr_image, boxplot_cjv_image],
                [boxplot_qi2_image],
            ]
            noise_table = Table(noise_data, [90 * mm, 70 * mm])
            noise_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(noise_table)
            self.report.append(PageBreak())

            # 4 page - Boxplots  #############################################
            ##################################################################
            # FBER, EFC
            self.report.append(
                Paragraph(
                    (
                        "<font size = 18 > <b>Spatial "
                        "distribution</b> </font>"
                    ),
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))
            boxplot_fber_image = Image(
                self.mriqc_group["FBER"], width=3 * inch, height=2.5 * inch
            )
            boxplot_efc_image = Image(
                self.mriqc_group["EFC"], width=3 * inch, height=2.5 * inch
            )
            spatial_distribution_data = [
                [boxplot_fber_image, boxplot_efc_image]
            ]
            spatial_distribution_table = Table(
                spatial_distribution_data, [90 * mm, 70 * mm]
            )
            spatial_distribution_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(spatial_distribution_table)
            self.report.append(Spacer(0 * mm, 4 * mm))
            self.report.append(Spacer(0 * mm, 4 * mm))

            # wm2max, inu_range, inu_med
            self.report.append(
                Paragraph(
                    "<font size = 18 > <b>ARTIFACTS</b> </font>",
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))
            boxplot_w2max_image = Image(
                self.mriqc_group["WM2MAX"], width=3 * inch, height=2.5 * inch
            )
            boxplot_inu_image = Image(
                self.mriqc_group["INU"], width=3 * inch, height=2.5 * inch
            )
            artifact_data = [[boxplot_w2max_image, boxplot_inu_image]]
            artifact_table = Table(artifact_data, [90 * mm, 70 * mm])
            artifact_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(artifact_table)
            self.report.append(PageBreak())

            # 4 page - Boxplots  #############################################
            ##################################################################
            self.report.append(
                Paragraph(
                    "<font size = 18 > <b>Tissues quality</b> </font>",
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))

            # FWHM, ICVS, RPVE, TMP_OVERLAP, SUMMARY
            boxplot_fwhm_image = Image(
                self.mriqc_group["FWHM"], width=3 * inch, height=2.5 * inch
            )
            boxplot_icvs_image = Image(
                self.mriqc_group["ICVS"], width=3 * inch, height=2.5 * inch
            )
            boxplot_rpve_image = Image(
                self.mriqc_group["RPVE"], width=3 * inch, height=2.5 * inch
            )
            boxplot_tmp_image = Image(
                self.mriqc_group["TMP_OVERLAP"],
                width=3 * inch,
                height=2.5 * inch,
            )
            boxplot_wm_image = Image(
                self.mriqc_group["SUMMARY_WM"],
                width=3 * inch,
                height=2.5 * inch,
            )
            boxplot_csf_image = Image(
                self.mriqc_group["SUMMARY_CSF"],
                width=3 * inch,
                height=2.5 * inch,
            )
            boxplot_gm_image = Image(
                self.mriqc_group["SUMMARY_GM"],
                width=3 * inch,
                height=2.5 * inch,
            )
            boxplot_bg_image = Image(
                self.mriqc_group["SUMMARY_BG"],
                width=3 * inch,
                height=2.5 * inch,
            )

            tissues_data = [
                [boxplot_fwhm_image, boxplot_icvs_image],
                [boxplot_rpve_image, boxplot_tmp_image],
                [boxplot_wm_image, boxplot_csf_image],
                [boxplot_gm_image, boxplot_bg_image],
            ]
            tissues_data_table = Table(tissues_data, [90 * mm, 70 * mm])
            tissues_data_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )

            self.report.append(tissues_data_table)

        elif modality == "bold":
            # Third page - Boxplots ##########################################
            ##################################################################
            self.report.append(
                Paragraph(
                    "<font size = 18 > <b>Artifact</b> </font>",
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))
            boxplot_aor_image = Image(
                self.mriqc_group["AOR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_aqi_image = Image(
                self.mriqc_group["AQI"], width=3 * inch, height=2.5 * inch
            )
            boxplot_dummy_image = Image(
                self.mriqc_group["DUMMY"], width=3 * inch, height=2.5 * inch
            )
            boxplot_gsr_image = Image(
                self.mriqc_group["GSR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_fd_image = Image(
                self.mriqc_group["FD"], width=3 * inch, height=2.5 * inch
            )
            boxplot_fd_num_image = Image(
                self.mriqc_group["FD_NUM"], width=3 * inch, height=2.5 * inch
            )
            boxplot_fd_perc_image = Image(
                self.mriqc_group["FD_PERC"], width=3 * inch, height=2.5 * inch
            )
            data = [
                [boxplot_aor_image, boxplot_aqi_image],
                [boxplot_dummy_image, boxplot_gsr_image],
                [boxplot_fd_image, boxplot_fd_num_image],
                [boxplot_fd_perc_image],
            ]
            table = Table(data, [90 * mm, 70 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(table)
            self.report.append(PageBreak())

            # 4 page - Boxplots ##############################################
            ##################################################################
            self.report.append(
                Paragraph(
                    (
                        "<font size = 18 > <b>Temporal "
                        "information</b> </font>"
                    ),
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))

            boxplot_dvars_image = Image(
                self.mriqc_group["DVARS"], width=3 * inch, height=2.5 * inch
            )
            boxplot_dvarsn_image = Image(
                self.mriqc_group["DVARSN"], width=3 * inch, height=2.5 * inch
            )
            boxplot_tsnr_num_image = Image(
                self.mriqc_group["TSNR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_gcor_perc_image = Image(
                self.mriqc_group["GCOR"], width=3 * inch, height=2.5 * inch
            )

            data = [
                [boxplot_dvars_image, boxplot_dvarsn_image],
                [boxplot_tsnr_num_image, boxplot_gcor_perc_image],
            ]
            table = Table(data, [90 * mm, 70 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(table)
            self.report.append(PageBreak())

            # 5 page - Boxplots ##############################################
            ##################################################################
            self.report.append(
                Paragraph(
                    (
                        "<font size = 18 > <b>Spatial "
                        "information</b> </font>"
                    ),
                    self.styles["Center"],
                )
            )
            self.report.append(Spacer(0 * mm, 4 * mm))
            line = ReportLine(150)
            line.hAlign = "CENTER"
            self.report.append(line)
            self.report.append(Spacer(0 * mm, 4 * mm))

            boxplot_fwhm_image = Image(
                self.mriqc_group["FWHM"], width=3 * inch, height=2.5 * inch
            )
            boxplot_icvs_image = Image(
                self.mriqc_group["SNR"], width=3 * inch, height=2.5 * inch
            )
            boxplot_rpve_image = Image(
                self.mriqc_group["FBER"], width=3 * inch, height=2.5 * inch
            )
            boxplot_tmp_image = Image(
                self.mriqc_group["EFC"], width=3 * inch, height=2.5 * inch
            )
            boxplot_gm_image = Image(
                self.mriqc_group["SUMMARY_FG"],
                width=3 * inch,
                height=2.5 * inch,
            )
            boxplot_bg_image = Image(
                self.mriqc_group["SUMMARY_BG"],
                width=3 * inch,
                height=2.5 * inch,
            )

            data = [
                [boxplot_fwhm_image, boxplot_icvs_image],
                [boxplot_rpve_image, boxplot_tmp_image],
                [boxplot_gm_image, boxplot_bg_image],
            ]
            table = Table(data, [90 * mm, 70 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            self.report.append(table)
            self.report.append(PageBreak())

        self.report.append(PageBreak())

        self.page.build(self.report, canvasmaker=PageNumCanvas)
